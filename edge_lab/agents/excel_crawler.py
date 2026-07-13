#!/usr/bin/env python3

import os
import csv
from pathlib import Path
from openpyxl import load_workbook


class UniversalExcelCrawler:
    def __init__(self, root_path, output_path):
        self.root_path = Path(root_path)
        self.output_path = Path(output_path)
        self.audit_data = []

    def get_file_size(self, file_path):
        return os.path.getsize(file_path)

    def get_file_size_human(self, bytes_size):
        for unit in ["B", "KB", "MB", "GB"]:
            if bytes_size < 1024.0:
                return f"{bytes_size:.2f} {unit}"
            bytes_size /= 1024.0
        return f"{bytes_size:.2f} TB"

    def crawl_excel_files(self):
        xlsx_count = 0
        xls_count = 0

        for file_path in self.root_path.rglob("*.xlsx"):
            xlsx_count += 1
            self._process_excel_file(file_path)

        for file_path in self.root_path.rglob("*.xls"):
            xls_count += 1
            self._process_excel_file(file_path)

        return xlsx_count, xls_count

    def _process_excel_file(self, file_path):
        try:
            file_size = self.get_file_size(file_path)
            relative_path = file_path.relative_to(self.root_path)

            wb = load_workbook(file_path, read_only=True, data_only=True)

            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                self.audit_data.append(
                    {
                        "filename": str(relative_path),
                        "file_size_bytes": file_size,
                        "file_size_human": self.get_file_size_human(file_size),
                        "sheet_name": sheet_name,
                        "sheet_index": sheet_idx,
                    }
                )

            wb.close()

        except Exception as e:
            print(f"Error processing {file_path}: {e}")

    def write_audit_index(self):
        os.makedirs(self.output_path.parent, exist_ok=True)

        fieldnames = [
            "filename",
            "file_size_bytes",
            "file_size_human",
            "sheet_name",
            "sheet_index",
        ]

        with open(self.output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.audit_data)

    def run(self):
        print(f"Starting crawl of: {self.root_path}")
        xlsx_count, xls_count = self.crawl_excel_files()
        print(f"Found {xlsx_count} .xlsx files and {xls_count} .xls files")
        print(f"Total sheets extracted: {len(self.audit_data)}")

        self.write_audit_index()
        print(f"Audit index written to: {self.output_path}")

        return len(self.audit_data)


if __name__ == "__main__":
    crawler = UniversalExcelCrawler(
        root_path="assets/charts/ОПР_статистика",
        output_path="data/audit_file_index.csv",
    )

    total_sheets = crawler.run()
    print(f"\nCrawl complete! Total sheets: {total_sheets}")
