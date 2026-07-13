#!/usr/bin/env python3
"""Crawler to map the Labor Market Excel file structure."""

import openpyxl
from pathlib import Path

FILE_PATH = Path(
    "/home/valalav/_projects/sirena-kbr/edge_lab/assets/charts/ОПР_статистика/ЗП_безработица/Зарплаты и СЧР (полная база).xlsx"
)


def map_sheets():
    """Map all sheets and find KBR data."""
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)

    print(f"Total sheets: {len(wb.sheetnames)}")
    print("\nSheet names:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"{i}. {name}")

    # Look for sheets containing KBR data
    kbr_sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Scan first 100 rows to find KBR
        for row_idx in range(1, min(101, ws.max_row + 1)):
            for col_idx in range(1, min(20, ws.max_column + 1)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and (
                    "Кабардино" in str(cell_value) or "Кбр" in str(cell_value)
                ):
                    kbr_sheets.append(
                        {
                            "sheet": sheet_name,
                            "row": row_idx,
                            "col": col_idx,
                            "value": str(cell_value)[:50],
                        }
                    )
                    break
            if any(s["sheet"] == sheet_name for s in kbr_sheets):
                break

    print(f"\nSheets containing KBR data: {len(kbr_sheets)}")
    for sheet_info in kbr_sheets:
        print(
            f"  - {sheet_info['sheet']}: Row {sheet_info['row']}, Col {sheet_info['col']}, '{sheet_info['value']}'"
        )

    wb.close()
    return kbr_sheets


if __name__ == "__main__":
    map_sheets()
