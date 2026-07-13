#!/usr/bin/env python3
"""
Task 120: Deep Dive - GRP Forecasts (15MB)

Extract historical and forecast GRP data for KBR from:
'Опережающий индикатор ВРП (с прогнозом).xlsm'

OUTPUT: data/kbr_grp_forecast.csv
"""

import pandas as pd
import openpyxl
from pathlib import Path
import datetime


def extract_kbr_grp_data():
    """Extract KBR GRP historical and forecast data."""

    # File path
    file_path = Path(
        "./assets/charts/ОПР_статистика/ВРП/Опережающий индикатор ВРП (с прогнозом).xlsm"
    )

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Load workbook
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Extract historical data
    print("Extracting historical GRP data...")
    ws_hist = wb["Данные_Регионы"]
    historical_data = []

    for row in ws_hist.iter_rows(min_row=2, values_only=True):
        region = row[5] if len(row) > 5 else None
        if region and "Кабардино" in str(region):
            date_str = row[0]
            data_type = row[1]
            value = row[2]
            year = row[3]
            quarter = row[4]
            date_ordinal = row[7] if len(row) > 7 else None

            historical_data.append(
                {
                    "date_str": date_str,
                    "type": data_type,
                    "value": value,
                    "year": year,
                    "quarter": quarter,
                    "date_ordinal": date_ordinal,
                    "is_forecast": False,
                }
            )

    # Extract forecast data
    print("Extracting forecast GRP data...")
    ws_fcst = wb["Данные_прогнозы"]
    forecast_data = []

    for row in ws_fcst.iter_rows(min_row=2, values_only=True):
        region = row[6] if len(row) > 6 else None
        if region and "Кабардино" in str(region):
            date_str = row[0]
            method = row[1]
            data_type = row[2]
            value = row[3]
            year = row[4]
            quarter = row[5]
            date_ordinal = row[8] if len(row) > 8 else None

            forecast_data.append(
                {
                    "date_str": date_str,
                    "method": method,
                    "type": data_type,
                    "value": value,
                    "year": year,
                    "quarter": quarter,
                    "date_ordinal": date_ordinal,
                    "is_forecast": True,
                }
            )

    # Create DataFrame for Base_SA only
    hist_df = pd.DataFrame([d for d in historical_data if d["type"] == "Base_SA"])
    fcst_df = pd.DataFrame([d for d in forecast_data if d["type"] == "Base_SA"])

    # Convert quarterly to monthly
    print("Converting quarterly to monthly...")

    def quarter_to_monthly(row):
        """Convert quarterly data to 3 monthly points."""
        year = row["year"]
        quarter = row["quarter"]

        # Map quarter to starting month
        quarter_map = {1: 1, 2: 4, 3: 7, 4: 10}
        start_month = quarter_map[quarter]

        months = []
        for i in range(3):
            month = start_month + i
            if month > 12:
                month -= 12
                year_adj = year + 1
            else:
                year_adj = year

            date_obj = datetime.datetime(year_adj, month, 1)
            months.append(
                {
                    "date": date_obj.strftime("%Y-%m-01"),
                    "year": year_adj,
                    "month": month,
                    "quarter": quarter,
                    "grp_index_base_sa": row["value"],
                    "is_forecast": row["is_forecast"],
                }
            )

        return months

    # Process historical data
    monthly_data = []
    for _, row in hist_df.iterrows():
        monthly_data.extend(quarter_to_monthly(row))

    # Process forecast data
    for _, row in fcst_df.iterrows():
        monthly_data.extend(quarter_to_monthly(row))

    # Create DataFrame and remove duplicates (forecasts override historical for same period)
    df = pd.DataFrame(monthly_data)
    df = df.sort_values(["year", "month"])

    # Drop duplicates (keep forecast if available)
    df = df.drop_duplicates(subset=["date"], keep="last")

    # Select and reorder columns
    result = df[
        ["date", "year", "month", "quarter", "grp_index_base_sa", "is_forecast"]
    ]

    # Calculate forecast horizon based on is_forecast flag
    result["forecast_horizon"] = result["is_forecast"].apply(
        lambda x: "Forecast" if x else "Historical"
    )

    print(f"\nTotal rows extracted: {len(result)}")
    print(f"Date range: {result['date'].iloc[0]} to {result['date'].iloc[-1]}")
    print(f"Historical rows: {len(result[result['is_forecast'] == False])}")
    print(f"Forecast rows: {len(result[result['is_forecast'] == True])}")

    # Save to CSV
    output_path = Path("data/kbr_grp_forecast.csv")
    output_path.parent.mkdir(exist_ok=True)
    result.to_csv(output_path, index=False)

    print(f"\n✓ Data saved to: {output_path}")

    return result


if __name__ == "__main__":
    df = extract_kbr_grp_data()
    print("\n=== Sample Data ===")
    print(df.head(10).to_string())
    print("\n=== Forecast Data ===")
    print(df[df["is_forecast"] == True].to_string())
