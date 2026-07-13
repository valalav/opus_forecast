#!/usr/bin/env python3
"""
Task 124: Mining: High-Freq Indicators (HH.ru & DomClick)

Extract HH Index and Housing Prices for KBR from OPR stats files.
Since DomClick Excel files use external data connections that are not accessible,
we extract housing prices from Rosstat data instead (from "Основная статистика ЮГУ.xlsx").

Outputs:
- data/kbr_high_freq_indicators.csv: Combined high-frequency indicators
- data/task124_report.md: Analysis report with correlation
"""

import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime
import json
import os


def parse_hh_index(
    file_path: str, target_region: str = "Кабардино-Балкарская Республика"
) -> pd.DataFrame:
    """Parse HH Index Excel file for labor market tension.

    The file is in long format with columns: name, name_prof, rep_date, HHI
    """
    df = pd.read_excel(file_path, sheet_name=0)

    # Filter for KBR
    df_kbr = df[
        df["name"].str.contains(target_region, na=False)
        | df["name"].str.contains("Кабардино", na=False)
    ].copy()

    if df_kbr.empty:
        raise ValueError(f"Could not find {target_region} in HH Index file")

    # Aggregate by date (average across all professions)
    df_kbr = df_kbr.groupby("rep_date")["HHI"].mean().reset_index()
    df_kbr.columns = ["Date", "HH_Index"]

    # Filter out empty values
    df_kbr = df_kbr[df_kbr["HH_Index"].notna() & (df_kbr["HH_Index"] != "")]

    # Convert to numeric and datetime
    df_kbr["HH_Index"] = pd.to_numeric(df_kbr["HH_Index"], errors="coerce")
    df_kbr["Date"] = pd.to_datetime(df_kbr["Date"])

    df_kbr = df_kbr.sort_values("Date").reset_index(drop=True)
    df_kbr["Date"] = df_kbr["Date"].dt.strftime("%Y-%m-%d")

    return df_kbr


def parse_housing_prices(file_path: str, target_region: str = "Кбр") -> pd.DataFrame:
    """Parse Housing Prices from Rosstat data (sheet 118 of YUGU stats)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "118" not in wb.sheetnames:
        raise ValueError("Sheet '118' not found in housing stats file")

    ws = wb["118"]

    data_secondary = []
    data_primary = []

    # Parse dates from row 11 (starting from column 2)
    dates = []
    for i, cell in enumerate(ws[11], 1):
        if i > 1 and isinstance(cell.value, datetime):
            dates.append(cell.value)

    # Find Kbr rows for secondary and primary markets
    secondary_kbr_row = None
    primary_kbr_row = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= 60:
            region_code = row[0] if row[0] else ""

            # Secondary market row ("вторичный рынок" above, Kbr row below)
            if region_code == target_region and row_idx > 20 and row_idx < 32:
                secondary_kbr_row = row
            # Primary market row
            if region_code == target_region and row_idx > 40 and row_idx < 50:
                primary_kbr_row = row

    if secondary_kbr_row is None:
        raise ValueError(
            f"Could not find secondary market housing data for {target_region}"
        )

    # Extract secondary market prices (starting from column 2)
    for i, (date, price) in enumerate(zip(dates, secondary_kbr_row[1:]), 1):
        if i <= len(dates) and isinstance(price, (int, float)) and not pd.isna(price):
            data_secondary.append(
                {"Date": date.strftime("%Y-%m-%d"), "Housing_Price_Secondary": price}
            )

    # Extract primary market prices
    if primary_kbr_row:
        for i, (date, price) in enumerate(zip(dates, primary_kbr_row[1:]), 1):
            if (
                i <= len(dates)
                and isinstance(price, (int, float))
                and not pd.isna(price)
            ):
                data_primary.append(
                    {"Date": date.strftime("%Y-%m-%d"), "Housing_Price_Primary": price}
                )

    df_secondary = pd.DataFrame(data_secondary)
    df_primary = pd.DataFrame(data_primary)

    if not df_secondary.empty:
        df_secondary["Date"] = pd.to_datetime(df_secondary["Date"])
        df_secondary = df_secondary.sort_values("Date").reset_index(drop=True)

    if not df_primary.empty:
        df_primary["Date"] = pd.to_datetime(df_primary["Date"])
        df_primary = df_primary.sort_values("Date").reset_index(drop=True)

    # Merge primary and secondary
    df = df_secondary.merge(df_primary, on="Date", how="outer")

    return df


def load_cpi_data(file_path: str = "data/infl_kbr.csv") -> pd.DataFrame:
    """Load CPI data for correlation analysis."""
    if not os.path.exists(file_path):
        return None

    df = pd.read_csv(file_path)
    if "date" in df.columns:
        df["Date"] = pd.to_datetime(df["date"])
    elif "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"])

    # Get monthly CPI change (mom)
    if "cpi" in df.columns:
        df["CPI_mom"] = df["cpi"]
    elif "mom" in df.columns:
        df["CPI_mom"] = df["mom"]

    return df[["Date", "CPI_mom"]].dropna() if "CPI_mom" in df.columns else None


def calculate_correlation(
    df1: pd.DataFrame, col1: str, df2: pd.DataFrame, col2: str, max_lag: int = 6
) -> dict:
    """Calculate cross-correlation between two time series."""
    merged = df1[["Date", col1]].merge(df2[["Date", col2]], on="Date", how="inner")

    if len(merged) < 6:
        return {"error": "Insufficient data points for correlation"}

    results = {}
    for lag in range(max_lag + 1):
        if lag == 0:
            corr = merged[col1].corr(merged[col2])
        else:
            series1 = merged[col1].iloc[:-lag].values
            series2 = merged[col2].iloc[lag:].values
            if len(series1) > 0 and len(series2) > 0:
                corr = np.corrcoef(series1, series2)[0, 1]
            else:
                corr = np.nan
        results[f"lag_{lag}"] = corr

    return results


def main():
    print("Task 124: Mining High-Freq Indicators (HH.ru & Housing Prices)")
    print("=" * 70)

    # Create output directory
    os.makedirs("data", exist_ok=True)

    # Parse HH Index
    print("\n1. Parsing HH Index...")
    hh_file = "data/raw/opr_stat/hh_индекс.xlsx"
    if os.path.exists(hh_file):
        try:
            df_hh = parse_hh_index(hh_file)
            print(f"   - Loaded {len(df_hh)} rows of HH Index data")
        except Exception as e:
            print(f"   - Error: {e}")
            df_hh = pd.DataFrame(columns=["Date", "HH_Index"])
    else:
        print(f"   - File not found: {hh_file}")
        df_hh = pd.DataFrame(columns=["Date", "HH_Index"])

    # Parse Housing Prices (using Rosstat data instead of DomClick)
    print("\n2. Parsing Housing Prices...")
    housing_file = "data/raw/opr_stat/Основная статистика ЮГУ.xlsx"
    if os.path.exists(housing_file):
        try:
            df_housing = parse_housing_prices(housing_file)
            print(f"   - Loaded {len(df_housing)} rows of housing price data")
            print(
                f"   - Secondary market prices: {df_housing['Housing_Price_Secondary'].notna().sum()} values"
            )
            print(
                f"   - Primary market prices: {df_housing['Housing_Price_Primary'].notna().sum()} values"
            )
        except Exception as e:
            print(f"   - Error: {e}")
            df_housing = pd.DataFrame(
                columns=["Date", "Housing_Price_Secondary", "Housing_Price_Primary"]
            )
    else:
        print(f"   - File not found: {housing_file}")
        df_housing = pd.DataFrame(
            columns=["Date", "Housing_Price_Secondary", "Housing_Price_Primary"]
        )

    # Load CPI data for correlation
    print("\n3. Loading CPI data...")
    df_cpi = load_cpi_data()
    if df_cpi is not None:
        print(f"   - Loaded {len(df_cpi)} rows of CPI data")
    else:
        print("   - CPI data not found")

    # Merge all data
    print("\n4. Merging indicators...")

    # Convert all Date columns to datetime
    df_hh["Date"] = pd.to_datetime(df_hh["Date"])
    df_housing["Date"] = pd.to_datetime(df_housing["Date"])
    if df_cpi is not None:
        df_cpi["Date"] = pd.to_datetime(df_cpi["Date"])

    df_merged = df_hh.merge(df_housing, on="Date", how="outer")

    if df_cpi is not None:
        df_merged = df_merged.merge(df_cpi, on="Date", how="outer")

    df_merged = df_merged.sort_values("Date").reset_index(drop=True)
    print(f"   - Merged dataset: {len(df_merged)} rows")

    # Save to CSV
    output_file = "data/kbr_high_freq_indicators.csv"
    df_merged.to_csv(output_file, index=False)
    print(f"   - Saved to: {output_file}")

    # Calculate correlations
    print("\n5. Calculating correlations...")
    correlations = {}

    if not df_hh.empty and df_cpi is not None:
        corr_hh_cpi = calculate_correlation(df_hh, "HH_Index", df_cpi, "CPI_mom")
        correlations["HH_Index vs CPI"] = corr_hh_cpi
        print(f"   - HH Index vs CPI: {corr_hh_cpi}")

    if not df_housing.empty and df_cpi is not None:
        df_housing_sec = df_housing[["Date", "Housing_Price_Secondary"]].dropna()
        corr_housing_cpi = calculate_correlation(
            df_housing_sec, "Housing_Price_Secondary", df_cpi, "CPI_mom"
        )
        correlations["Housing_Price vs CPI"] = corr_housing_cpi
        print(f"   - Housing Price vs CPI: {corr_housing_cpi}")

    if not df_hh.empty and not df_housing.empty:
        df_housing_sec = df_housing[["Date", "Housing_Price_Secondary"]].dropna()
        corr_hh_housing = calculate_correlation(
            df_hh, "HH_Index", df_housing_sec, "Housing_Price_Secondary"
        )
        correlations["HH_Index vs Housing_Price"] = corr_hh_housing
        print(f"   - HH Index vs Housing Price: {corr_hh_housing}")

    # Generate report
    print("\n6. Generating report...")
    report = f"""# Task 124: High-Frequency Indicators Extraction Report

**Date:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

## Overview

This report documents the extraction of high-frequency leading indicators for KBR:
1. **HH Index** (hh.ru): Labor market tension indicator
2. **Housing Prices** (Rosstat): Primary and secondary market prices (quarterly data)

## Data Sources

| Source | File | Type | Frequency |
|--------|------|------|-----------|
| HH Index | `data/raw/opr_stat/hh_индекс.xlsx` | Labor market tension | Monthly |
| Housing Prices | `data/raw/opr_stat/Основная статистика ЮГУ.xlsx` (sheet 118) | Rosstat housing prices | Quarterly |

## Note on DomClick Data

The original task specified parsing DomClick housing price data from:
- `Цены на жилье (Домклик, факт.сделки).xlsx`
- `Цены на жилье (Домклик, обьявления).xlsx`

These Excel files use **external Power Query data connections** to the SberIndex portal. The raw data is NOT stored in the Excel file itself - it is fetched dynamically via API when the dashboard is opened in Excel with a live connection. This technical limitation prevents extracting KBR-specific data using standard Excel parsing libraries (openpyxl, pandas, etc.).

**Alternative approach:** Housing price data for KBR was extracted from Rosstat statistics instead, which provides quarterly housing price data for both primary and secondary markets.

## Extracted Data

- **HH Index**: {len(df_hh)} data points (monthly)
- **Housing Prices**: {len(df_housing)} data points (quarterly)
- **Output file**: `data/kbr_high_freq_indicators.csv`

## Correlation Analysis

### HH Index vs CPI

```json
{json.dumps(correlations.get("HH_Index vs CPI", {}), indent=2)}
```

### Housing Prices vs CPI

```json
{json.dumps(correlations.get("Housing_Price vs CPI", {}), indent=2)}
```

### HH Index vs Housing Prices

```json
{json.dumps(correlations.get("HH_Index vs Housing_Price", {}), indent=2)}
```

## Data Quality Assessment

| Indicator | Data Points | Missing Values | Frequency | Notes |
|-----------|-------------|----------------|-----------|-------|
| HH_Index | {len(df_hh)} | {df_hh["HH_Index"].isna().sum()} | Monthly | High-frequency labor market indicator |
| Housing_Price_Secondary | {df_housing["Housing_Price_Secondary"].notna().sum()} | {df_housing["Housing_Price_Secondary"].isna().sum()} | Quarterly | Rosstat secondary market prices |
| Housing_Price_Primary | {df_housing["Housing_Price_Primary"].notna().sum()} | {df_housing["Housing_Price_Primary"].isna().sum()} | Quarterly | Rosstat primary market prices |

## Recommendations

1. **For real-time housing price data**: Consider establishing a direct API connection to SberIndex portal if the API endpoints become accessible.

2. **For forecasting use**:
   - HH Index: High-frequency monthly indicator (use with 1-3 month lag for leading indicator)
   - Housing Prices: Quarterly indicator (use with 3-6 month lag for cost-push inflation signals)

3. **Integration**: These indicators can be added to the RidgeMacroForecaster as additional exogenous variables.

## Sample Output (Last 5 rows)

```
{df_merged.tail().to_string(index=False)}
```
"""

    report_file = "data/task124_report.md"
    with open(report_file, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"   - Report saved to: {report_file}")

    print("\n" + "=" * 70)
    print("Task 124 completed successfully!")
    print(f"Output: {output_file}")
    print(f"Report: {report_file}")

    return df_merged, correlations


if __name__ == "__main__":
    df_merged, correlations = main()
