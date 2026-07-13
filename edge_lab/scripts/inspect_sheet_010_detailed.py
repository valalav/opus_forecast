import openpyxl

wb_path = "/home/valalav/_projects/sirena-kbr/edge_lab/assets/charts/ОПР_статистика/Основная статистика ЮГУ.xlsx"
wb = openpyxl.load_workbook(wb_path)
ws = wb["010"]

print("=== Row 10 ===")
row_10 = []
for col_idx in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=10, column=col_idx).value
    row_10.append(str(cell_value)[:20] if cell_value else "")
print(f"Columns 1-30: {row_10[:30]}")

print("\n=== Row 11 (Header Row) ===")
row_11 = []
for col_idx in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=11, column=col_idx).value
    row_11.append(str(cell_value)[:20] if cell_value else "")
print(f"Columns 1-30: {row_11[:30]}")

print("\n=== Row 12 (First data row) ===")
row_12 = []
for col_idx in range(1, min(50, ws.max_column + 1)):
    cell_value = ws.cell(row=12, column=col_idx).value
    row_12.append(str(cell_value)[:10] if cell_value else "")
print(f"Columns 1-50: {row_12}")

print("\n=== Row 18 (First row with data) ===")
row_18 = []
for col_idx in range(1, min(50, ws.max_column + 1)):
    cell_value = ws.cell(row=18, column=col_idx).value
    row_18.append(str(cell_value)[:10] if cell_value else "")
print(f"Columns 1-50: {row_18}")

print(f"\nTotal rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")
