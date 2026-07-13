#!/usr/bin/env python3
"""
Extract macro-economic series from Sheet '010' of 'Основная статистика ЮГУ.xlsx'
"""

import pandas as pd
import openpyxl
from datetime import datetime

wb_path = "/home/valalav/_projects/sirena-kbr/edge_lab/assets/charts/ОПР_статистика/Основная статистика ЮГУ.xlsx"
output_path = "/home/valalav/_projects/sirena-kbr/edge_lab/data/kbr_macro_monolith.csv"

wb = openpyxl.load_workbook(wb_path, data_only=True)
ws = wb["010"]

# Verify region 'Кбр' at cell (7, 2)
region_check = ws.cell(row=7, column=2).value
print(f"Region check (row 7, col 2): {region_check}")
if "Кбр" not in str(region_check):
    print(f"WARNING: Expected 'Кбр' but found '{region_check}'")

# Extract dates from row 11 (columns 5 onwards are dates)
dates = []
for col_idx in range(5, ws.max_column + 1):
    cell_value = ws.cell(row=11, column=col_idx).value
    if cell_value and isinstance(cell_value, datetime):
        dates.append(cell_value.strftime("%Y-%m-%d"))
    elif cell_value:
        dates.append(str(cell_value))

print(f"Found {len(dates)} date columns")

# Iterate rows from 12 to EOF
output_data = []
unique_series = set()

for row_idx in range(12, ws.max_row + 1):
    indicator = ws.cell(row=row_idx, column=1).value
    category = ws.cell(row=row_idx, column=2).value
    metric_type = ws.cell(row=row_idx, column=4).value

    # Skip rows without indicator
    if not indicator:
        continue

    indicator_str = str(indicator).strip()
    category_str = str(category).strip() if category else ""
    metric_type_str = str(metric_type).strip() if metric_type else ""

    # Extract values for each date (starting at column 5)
    for date_idx, date_str in enumerate(dates):
        col_idx = 5 + date_idx
        value = ws.cell(row=row_idx, column=col_idx).value

        if value is not None and value != "":
            output_data.append(
                {
                    "Date": date_str,
                    "Indicator": indicator_str,
                    "Category": category_str,
                    "Metric_Type": metric_type_str,
                    "Value": float(value) if isinstance(value, (int, float)) else value,
                }
            )

            # Track unique series
            series_key = f"{indicator_str}::{category_str}::{metric_type_str}"
            unique_series.add(series_key)

print(f"Total records: {len(output_data)}")
print(f"Unique series: {len(unique_series)}")

# Write to CSV
if output_data:
    df = pd.DataFrame(output_data)
    df.to_csv(output_path, index=False)
    print(f"Data written to {output_path}")

    # Print summary
    print(f"\nTop 10 indicators:")
    print(df["Indicator"].value_counts().head(10))

    print(f"\nMetric types:")
    print(df["Metric_Type"].value_counts())

    # Date range
    print(f"\nDate range: {df['Date'].min()} to {df['Date'].max()}")
else:
    print("No data extracted!")
