#!/usr/bin/env python3
"""Deep dive into Labor Market Excel file structure."""

import openpyxl
from pathlib import Path

FILE_PATH = Path(
    "/home/valalav/_projects/sirena-kbr/edge_lab/assets/charts/ОПР_статистика/ЗП_безработица/Зарплаты и СЧР (полная база).xlsx"
)


def explore_sheet(sheet_name, max_rows=200):
    """Explore a specific sheet to understand its structure."""
    print(f"\n{'=' * 60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 60}")

    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb[sheet_name]

    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")

    # Print first 30 rows to understand structure
    print(f"\nFirst 30 rows (first 15 columns):")
    for row_idx in range(1, min(31, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(16, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            val = cell.value
            if val is not None:
                row_data.append(str(val)[:30])
            else:
                row_data.append("")

        # Only print non-empty rows
        if any(row_data):
            print(f"Row {row_idx:3d}: {' | '.join(row_data)}")

    # Search for KBR-related patterns
    print(f"\nSearching for KBR patterns in first 200 rows...")
    patterns = [
        "Кабардино",
        "КБР",
        "Кабардино-Балкар",
        "Северный Кавказ",
        "СКФО",
        "Кавказ",
        "Республика",
        "край",
        "область",
    ]

    for pattern in patterns:
        for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
            for col_idx in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value and pattern.lower() in str(cell.value).lower():
                    print(
                        f"  Found '{pattern}' at Row {row_idx}, Col {col_idx}: {str(cell.value)[:60]}"
                    )

    wb.close()


def main():
    """Main exploration."""
    # Focus on data sheets
    data_sheets = ["Номиналы", "Динамика", "Отрасли_для_графиков"]

    for sheet in data_sheets:
        explore_sheet(sheet)


if __name__ == "__main__":
    main()
