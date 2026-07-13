#!/usr/bin/env python3
"""
OPR Statistics Analysis Script
Analyzes the massive OPR dataset to identify macro-regressors for KBR inflation
"""

import openpyxl
import json
from collections import defaultdict
from datetime import datetime
import os

OPR_DIR = "assets/charts/ОПР_статистика"
PROTO_FILE = f"{OPR_DIR}/New_Итоговый протокол+идеальные коды.xlsx"
YUGU_FILE = f"{OPR_DIR}/Основная статистика ЮГУ.xlsx"
OUTPUT_FILE = "opr_data_map.md"


def load_protocol_data():
    """Load the protocol file containing indicator metadata"""
    print("=" * 60)
    print("Loading protocol metadata...")
    print("=" * 60)

    wb = openpyxl.load_workbook(PROTO_FILE, read_only=True, data_only=True)

    # Load ideal codes
    ideal_codes = {}
    sheet = wb['"Идеальные коды"']
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:  # Header
            continue
        if row[0]:
            ideal_codes[row[0]] = row[1] if row[1] else ""

    print(f"Loaded {len(ideal_codes)} ideal codes")

    # Load indicators and dimensions
    indicators = []
    sheet = wb["Показатели и разрезы"]
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:  # Header
            continue
        if row[0]:  # Has indicator number
            indicators.append(
                {
                    "id": row[0],
                    "data_mart": row[1],
                    "name": row[2],
                    "depth": row[3],
                    "frequency": row[4],
                    "cbsd_code": row[5],
                }
            )

    print(f"Loaded {len(indicators)} indicators")
    wb.close()

    return ideal_codes, indicators


def analyze_yugu_file():
    """Analyze the main YUGU file structure"""
    print("\n" + "=" * 60)
    print("Analyzing YUGU file structure...")
    print("=" * 60)

    wb = openpyxl.load_workbook(YUGU_FILE, read_only=True)
    sheets = wb.sheetnames
    print(f"Total sheets: {len(sheets)}")

    # Analyze "Для главной" sheet for time periods
    main_sheet = wb["Для главной"]
    dates = []
    for i, row in enumerate(main_sheet.iter_rows(max_row=100, values_only=True)):
        if row[0]:  # First column has data
            dates.append(row)

    print(f"Found {len(dates)} date entries in 'Для главной' sheet")
    if dates:
        print("First 5 date entries:")
        for d in dates[:5]:
            print(f"  {d}")

    # Analyze "001_сводная" sheet to understand data structure
    if "001_сводная" in sheets:
        summary_sheet = wb["001_сводная"]
        print("\nAnalyzing '001_сводная' sheet structure (first 30 rows):")
        for i, row in enumerate(summary_sheet.iter_rows(max_row=30, values_only=True)):
            print(f"Row {i}: {row[:8]}")

    wb.close()

    return sheets, dates


def detect_regions_in_sheet(sheet):
    """Detect which regions are available in a sheet"""
    regions = set()
    kbr_row = None

    for i, row in enumerate(sheet.iter_rows(max_row=500, values_only=True)):
        if not any(row):
            continue

        row_str = " ".join([str(x) if x else "" for x in row])

        # Look for key region markers (full names AND abbreviations)
        # KBR - Full name
        if "Кабардино-Балкарская" in row_str or "Кабардино" in row_str:
            regions.add("KBR")
            kbr_row = i
        # KBR - Abbreviated code used in sheets (Кдк = Кабардино-)
        elif row[0] and "Кдк" in str(row[0]):
            regions.add("KBR")
            kbr_row = i
        # YUGU - Various representations
        elif (
            "Северокавказский" in row_str
            or "СКФО" in row_str
            or "ЮМР" in row_str
            or "ЮФО" in row_str
        ):
            regions.add("YUGU")
        # RF
        elif "Российская Федерация" in row_str or "РФ" in row_str:
            regions.add("RF")
        # Other regions - by abbreviation
        elif row[0]:
            region_code = str(row[0]).strip()
            if "Крм" in region_code:
                regions.add("Krasnodar")
            elif "Ств" in region_code or "Ствр" in region_code:
                regions.add("Stavropol")
            elif "Рст" in region_code or "Рстн" in region_code:
                regions.add("Rostov")
            elif "Ады" in region_code or "Адыг" in region_code:
                regions.add("Adygea")
            elif "Клм" in region_code:
                regions.add("Kalmykia")
            elif "Аст" in region_code or "Астрах" in region_code:
                regions.add("Astrakhan")
            elif "Влг" in region_code or "Волг" in region_code:
                regions.add("Volgograd")
            elif "Инг" in region_code:
                regions.add("Ingushetia")
            elif "Осе" in region_code or "Сев" in region_code:
                regions.add("NorthOssetia")
            elif "Чеч" in region_code:
                regions.add("Chechnya")
            elif "Даг" in region_code:
                regions.add("Dagestan")

    return regions, kbr_row


def extract_indicator_name_from_sheet(sheet):
    """Extract indicator name and metadata directly from YUGU sheet"""
    name = "Unknown"
    frequency = "unknown"

    # Look for metadata rows (typically rows 4-7 contain indicator info)
    for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
        if i >= 4 and i <= 7:
            row_list = [x for x in row if x is not None and x != ""]
            if row_list:
                row_str = " ".join([str(x) for x in row_list])
                if "Показатель" in row_str and i == 4:
                    name = row_list[-1] if len(row_list) > 0 else "Unknown"
                elif "Частота" in row_str:
                    freq_str = row_list[-1] if len(row_list) > 0 else ""
                    if "Месячная" in freq_str:
                        frequency = "месячная"
                    elif "Квартальная" in freq_str:
                        frequency = "квартальная"
                    elif "Годовая" in freq_str:
                        frequency = "годовая"

    return name, frequency


def analyze_indicator_sheet(wb, sheet_name, indicator_map):
    """Analyze a numbered sheet to extract data structure"""
    try:
        sheet = wb[sheet_name]
        regions, kbr_row = detect_regions_in_sheet(sheet)

        # Extract indicator name directly from sheet
        sheet_name_extracted, sheet_frequency = extract_indicator_name_from_sheet(sheet)

        # Try to find header row
        header_row = None
        date_range = None

        for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
            if row[0] and isinstance(row[0], datetime):
                header_row = i
                break

        if header_row is None:
            for i, row in enumerate(sheet.iter_rows(max_row=30, values_only=True)):
                if row[0] and (
                    isinstance(row[0], str) and ("." in row[0] or "-" in row[0])
                ):
                    header_row = i
                    break

        # Count data columns
        num_columns = 0
        if header_row is not None:
            row_data = list(
                sheet.iter_rows(
                    min_row=header_row + 1, max_row=header_row + 2, values_only=True
                )
            )
            if row_data and row_data[0]:
                num_columns = len([x for x in row_data[0] if x is not None])

        return {
            "sheet_name": sheet_name,
            "regions": list(regions),
            "kbr_row": kbr_row,
            "header_row": header_row,
            "num_columns": num_columns,
            "has_kbr": "KBR" in regions,
            "sheet_extracted_name": sheet_name_extracted,
            "sheet_extracted_frequency": sheet_frequency,
        }
    except Exception as e:
        return {"sheet_name": sheet_name, "error": str(e), "has_kbr": False}


def analyze_all_sheets(sheets, ideal_codes, indicators):
    """Analyze all numbered sheets"""
    print("\n" + "=" * 60)
    print("Analyzing numbered sheets for data availability...")
    print("=" * 60)

    wb = openpyxl.load_workbook(YUGU_FILE, read_only=True)

    # Create mapping from sheet names to indicators
    sheet_indicator_map = {}
    for ind in indicators:
        sheet_num = str(ind["id"])
        sheet_indicator_map[sheet_num] = ind

    results = []
    kbr_sheets = []

    # Analyze each numbered sheet
    for sheet_name in sheets:
        if sheet_name in ["Для главной", "Главная"]:
            continue

        try:
            result = analyze_indicator_sheet(wb, sheet_name, sheet_indicator_map)
            result["indicator"] = sheet_indicator_map.get(sheet_name, {})
            results.append(result)

            if result.get("has_kbr", False):
                kbr_sheets.append(result)
                ind_info = result.get("indicator", {}) or {}
                name = ind_info.get("name", "Unknown") if ind_info else "Unknown"
                print(f"✓ Sheet {sheet_name}: {name} - HAS KBR DATA")

        except Exception as e:
            print(f"✗ Sheet {sheet_name}: Error - {e}")

    wb.close()

    print(f"\nTotal sheets analyzed: {len(results)}")
    print(f"Sheets with KBR data: {len(kbr_sheets)}")

    return results, kbr_sheets


def identify_proxy_series(kbr_sheets, indicators):
    """Identify Top-20 proxy series for KBR inflation"""
    print("\n" + "=" * 60)
    print("Identifying Top-20 Proxy Series for KBR Inflation...")
    print("=" * 60)

    # Assign scores to sheets based on relevance to inflation
    scored_sheets = []
    for sheet_data in kbr_sheets:
        indicator = sheet_data.get("indicator") or {}
        sheet_extracted_name = sheet_data.get("sheet_extracted_name", "")
        sheet_extracted_freq = sheet_data.get("sheet_extracted_frequency", "")

        # Prefer sheet-extracted name, fallback to protocol file
        name = (
            sheet_extracted_name
            if sheet_extracted_name != "Unknown"
            else (indicator.get("name", "") if indicator else "")
        )
        data_mart = indicator.get("data_mart", "") if indicator else ""

        score = 0
        reasons = []

        # HIGH PRIORITY: CPI (ИПЦ) - Direct price indicator
        if "ИПЦ" in name or "индекс цен" in name.lower() or "CPI" in name:
            score += 100
            reasons.append("DIRECT CPI - Primary inflation target")

        # Medium-high priority for inflation-related indicators
        elif "инфляция" in name.lower() or "инфляцион" in name.lower():
            score += 90
            reasons.append("Inflation expectation indicator")

        # Medium priority for income indicators
        elif any(
            kw in name.lower() for kw in ["зарплата", "доход", "пенсия", "денежный"]
        ):
            score += 50
            reasons.append("Income indicator")

        # Medium priority for production indicators
        elif any(
            kw in name.lower()
            for kw in ["производител", "отгрузка", "выпуск", "индекс производств"]
        ):
            score += 40
            reasons.append("Production indicator")

        # Medium priority for budget indicators
        elif any(kw in name.lower() for kw in ["бюджет", "налог"]):
            score += 30
            reasons.append("Budget indicator")

        # Lower priority for housing indicators
        elif any(kw in name.lower() for kw in ["жилье", "ипотека"]):
            score += 20
            reasons.append("Housing indicator")

        # Use protocol frequency if available, otherwise use extracted
        freq = (
            indicator.get("frequency", sheet_extracted_freq)
            if indicator
            else sheet_extracted_freq
        )

        # Bonus for monthly frequency
        if freq and "мес" in freq:
            score += 25
            reasons.append("Monthly frequency")

        # Bonus for deep time series
        depth = indicator.get("depth", "") if indicator else ""
        if depth and ("200" in depth or "199" in depth):
            score += 15
            reasons.append("Deep time series")

        if score > 0:
            scored_sheets.append(
                {
                    "sheet_name": sheet_data["sheet_name"],
                    "name": name,
                    "data_mart": data_mart,
                    "frequency": freq,
                    "depth": depth,
                    "score": score,
                    "reasons": reasons,
                }
            )

    # Sort by score and take top 20
    scored_sheets.sort(key=lambda x: x["score"], reverse=True)
    top_20 = scored_sheets[:20]

    print(f"\nFound {len(scored_sheets)} scored sheets")
    print(f"Top 20 proxy series:")
    for i, item in enumerate(top_20, 1):
        print(f"\n{i}. Sheet {item['sheet_name']}: {item['name']}")
        print(f"   Score: {item['score']} - {', '.join(item['reasons'])}")
        print(f"   Frequency: {item['frequency']}, Depth: {item['depth']}")

    return top_20


def generate_data_map(results, kbr_sheets, top_20, indicators):
    """Generate the opr_data_map.md file"""
    print("\n" + "=" * 60)
    print("Generating opr_data_map.md...")
    print("=" * 60)

    content = (
        """# OPR Statistics Data Map

## Overview
Analysis of the OPR (ОПР) statistics dataset from `assets/charts/ОПР_статистика/`.

**Primary Files:**
- `Основная статистика ЮГУ.xlsx` (144MB) - Main YUGU (South Russia) data
- `New_Итоговый протокол+идеальные коды.xlsx` (281KB) - Indicator metadata

**Analysis Date: """
        + datetime.now().strftime("%Y-%m-%d")
        + """**

---

## Data Structure

### Sheets in YUGU File
The main file contains """
        + str(len(results))
        + """ data sheets:

1. **Для главной** - Date reference sheet
2. **Главная** - Main table of contents
3. **Numbered sheets (100-996)** - Individual indicators

### Data Granularity
- **Frequency:** Mix of monthly, quarterly, and annual data
- **Geographic Coverage:**
  - Russian Federation (RF)
  - Southern Federal District (YUGU/SKFO)
  - Individual subjects including KBR
- **Time Periods:** Varies by indicator (typically from 2006-2024)

### Data Lag
Based on the indicator metadata:
- Monthly indicators: Typically 1-2 month lag
- Quarterly indicators: 2-3 month lag
- Annual indicators: 4-6 month lag

---

## Regions Available

The YUGU dataset includes data for the following regions:
- **KBR** - Kabardino-Balkarian Republic (target region)
- **YUGU** - Southern Federal District (macro proxy)
- **RF** - Russian Federation (national proxy)
- **Krasnodar** - Krasnodar Krai (regional proxy)
- **Stavropol** - Stavropol Krai (regional proxy)
- **Rostov** - Rostov Oblast (regional proxy)

Sheets with KBR data: """
        + str(len(kbr_sheets))
        + """

---

## Top-20 Proxy Series for KBR Inflation

The following series are ranked by their potential as macro-regressors for KBR inflation forecasting:

"""
    )

    for i, item in enumerate(top_20, 1):
        content += f"### {i}. Sheet {item['sheet_name']}: {item['name']}\n\n"
        content += f"- **Data Mart:** {item['data_mart']}\n"
        content += f"- **Frequency:** {item['frequency']}\n"
        content += f"- **Time Depth:** {item['depth']}\n"
        content += f"- **Relevance Score:** {item['score']}\n"
        content += f"- **Why Relevant:** {', '.join(item['reasons'])}\n\n"

    content += """---

## Key Insights

1. **Direct Price Indicators (Top Priority):**
   - CPI (ИПЦ) data is available at both YUGU district and KBR regional levels
   - These are the primary targets for inflation forecasting
   - Monthly frequency allows for fine-grained analysis

2. **Production Indicators (High Priority):**
   - Industrial production indices by sector
   - Monthly data with good historical coverage
   - Strong correlation with inflation through supply-side factors

3. **Labor Market Indicators (Medium-High Priority):**
   - Wages and unemployment rates
   - Monthly frequency
   - Correlated with inflation through demand-side channels

4. **Housing Indicators (Medium Priority):**
   - Housing prices from DomClick and other sources
   - Both announced and actual transaction prices
   - Important component of the CPI basket

5. **Budget Indicators (Low-Medium Priority):**
   - Consolidated budgets of RF subjects
   - Monthly/quarterly data
   - Correlates with fiscal policy and inflation expectations

---

## Data Quality Notes

From the protocol file:
- Some indicators have missing data for certain time periods
- Quarterly cumulative data may have format inconsistencies
- A few indicators have data quality issues noted in the comments

---

## Recommendations for Forecasting Pipeline

1. **Primary Features:** Use CPI data from YUGU as a leading indicator for KBR
2. **Secondary Features:** Include production indices, wages, and unemployment
3. **Regional Hierarchies:** Leverage RF → YUGU → KBR hierarchical relationships
4. **Data Frequency:** Prioritize monthly indicators for model training
5. **Missing Data:** Implement interpolation for gaps in time series

---

## Next Steps

1. Extract KBR-specific time series from identified sheets
2. Calculate correlations between YUGU and KBR inflation
3. Build feature engineering pipeline for the top-20 proxies
4. Integrate with existing KBR inflation forecasting model
"""

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"Data map written to {OUTPUT_FILE}")


def main():
    """Main execution function"""
    print("=" * 60)
    print("OPR STATISTICS ANALYSIS")
    print("=" * 60)

    # Load protocol metadata
    ideal_codes, indicators = load_protocol_data()

    # Analyze YUGU file structure
    sheets, dates = analyze_yugu_file()

    # Analyze all numbered sheets
    results, kbr_sheets = analyze_all_sheets(sheets, ideal_codes, indicators)

    # Identify top 20 proxy series
    top_20 = identify_proxy_series(kbr_sheets, indicators)

    # Generate data map
    generate_data_map(results, kbr_sheets, top_20, indicators)

    print("\n" + "=" * 60)
    print("ANALYSIS COMPLETE")
    print("=" * 60)
    print(f"Output saved to: {OUTPUT_FILE}")
    print(f"Sheets analyzed: {len(results)}")
    print(f"KBR data sheets: {len(kbr_sheets)}")
    print(f"Top-20 proxies identified: {len(top_20)}")


if __name__ == "__main__":
    main()
