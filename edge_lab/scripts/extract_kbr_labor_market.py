#!/usr/bin/env python3
"""Extract KBR Labor Market data from the 111MB Excel file.

This script extracts:
- НЗП (Номинальная Заработная Плата) - Nominal Wage
- РЗП (Реальная Заработная Плата) - Real Wage
- ССЧР (Среднесписочная Численность Работников) - Employment

for KBR and outputs to CSV.
"""

import openpyxl
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

FILE_PATH = Path(
    "/home/valalav/_projects/sirena-kbr/edge_lab/assets/charts/ОПР_статистика/ЗП_безработица/Зарплаты и СЧР (полная база).xlsx"
)
OUTPUT_PATH = Path("/home/valalav/_projects/sirena-kbr/edge_lab/data/kbr_labor_market.csv")


def extract_from_report_sheet() -> pd.DataFrame:
    """Process Отчет sheet - main time series data source."""
    print("Processing Отчет sheet (main data source)...")

    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["Отчет"]

    data_rows = []

    # Get date columns from row 9 (columns 2+)
    date_cols = []
    for col in range(2, ws.max_column + 1):
        date_val = ws.cell(9, col).value
        if date_val and isinstance(date_val, datetime):
            date_cols.append((col, date_val))

    print(f"  Found {len(date_cols)} date columns")

    if not date_cols:
        wb.close()
        return pd.DataFrame()

    # Find all indicator rows (НЗП, РЗП, ССЧР)
    for row_idx in range(1, ws.max_row + 1):
        col1 = ws.cell(row_idx, 1).value

        # Check for indicator rows
        if col1 in ["НЗП", "РЗП", "ССЧР"]:
            indicator = col1

            # Extract industry data from subsequent rows
            # Each industry spans multiple rows (e.g., A. 02, A. 03, A. С/х, etc.)
            continue

        # Check for industry rows (starts with letter or "A.", "B.", etc.)
        elif col1 and isinstance(col1, str):
            # This is an industry row
            industry_code = col1

            # Find the indicator for this industry (look up to find НЗП/РЗП/ССЧР)
            indicator = None
            indicator_row = row_idx
            while indicator_row > 1:
                indicator_cell = ws.cell(indicator_row, 1).value
                if indicator_cell in ["НЗП", "РЗП", "ССЧР"]:
                    indicator = indicator_cell
                    break
                indicator_row -= 1

            # Skip if no indicator found
            if not indicator:
                continue

            # Extract values from all date columns
            for col_idx, date_val in date_cols:
                val = ws.cell(row_idx, col_idx).value
                if val is not None and val != "#N/A":
                    try:
                        numeric_val = (
                            float(val)
                            if isinstance(val, (int, float))
                            else float(str(val))
                        )
                        data_rows.append(
                            {
                                "Date": date_val.strftime("%Y-%m-%d"),
                                "Series_Name": f"{indicator}_{industry_code}",
                                "Indicator_Type": indicator,
                                "Industry_Code": industry_code,
                                "Region_Code": "KBR",
                                "Value": numeric_val,
                            }
                        )
                    except (ValueError, TypeError):
                        pass

    wb.close()

    if data_rows:
        df = pd.DataFrame(data_rows)
        df["Date"] = pd.to_datetime(df["Date"])
        df = df.sort_values(["Series_Name", "Date"]).reset_index(drop=True)
        print(f"  Extracted {len(df)} records")
        return df
    else:
        return pd.DataFrame()


def generate_summary(df: pd.DataFrame):
    """Generate summary statistics."""
    print("\n" + "=" * 60)
    print("SERIES SUMMARY")
    print("=" * 60)

    if df.empty:
        print("No data found!")
        return

    # Count unique series
    unique_series = df["Series_Name"].unique()
    print(f"Total unique series: {len(unique_series)}")

    # Count by indicator type
    print("\nSeries by indicator type:")
    for indicator in ["НЗП", "РЗП", "ССЧР"]:
        count = len([s for s in unique_series if indicator in str(s)])
        print(f"  {indicator} (Nominal/Real/Employment): {count} series")

    # Show sample series
    print("\nSample series (first 30):")
    for i, series in enumerate(sorted(unique_series)[:30], 1):
        count = len(df[df["Series_Name"] == series])
        print(f"  {i}. {series}: {count} records")

    if len(unique_series) > 30:
        print(f"  ... and {len(unique_series) - 30} more")

    # Date range
    print(f"\nDate range: {df['Date'].min()} to {df['Date'].max()}")
    print(f"Total records: {len(df)}")

    # Check acceptance criteria
    print("\n" + "=" * 60)
    print("ACCEPTANCE CRITERIA CHECK")
    print("=" * 60)

    # 1. Extracted > 50 labor market series
    print(
        f"1. Extracted > 50 labor market series: {'PASS' if len(unique_series) > 50 else 'FAIL'} ({len(unique_series)} series)"
    )

    # 2. Includes Nominal Wage (НЗП)
    has_nzp = any("НЗП" in str(s) for s in unique_series)
    print(f"2. Includes Nominal Wage (НЗП): {'PASS' if has_nzp else 'FAIL'}")

    # 3. Includes Real Wage (РЗП)
    has_rzp = any("РЗП" in str(s) for s in unique_series)
    print(f"3. Includes Real Wage (РЗП): {'PASS' if has_rzp else 'FAIL'}")

    # 4. Includes Employment (ССЧР)
    has_sschr = any("ССЧР" in str(s) for s in unique_series)
    print(f"4. Includes Employment (ССЧР): {'PASS' if has_sschr else 'FAIL'}")

    # 5. Includes sectoral breakdown
    has_industry = "Industry_Code" in df.columns and len(unique_series) > 30
    print(
        f"5. Includes sectoral breakdown (Wages by industry): {'PASS' if has_industry else 'FAIL'}"
    )


def pivot_to_wide_format(df: pd.DataFrame) -> pd.DataFrame:
    """Convert long format to wide format with series as columns."""
    if df.empty:
        return df

    # Pivot
    pivoted = df.pivot(index="Date", columns="Series_Name", values="Value")

    # Reset index
    pivoted = pivoted.reset_index()

    return pivoted


def main():
    """Main extraction function."""
    print(f"Extracting KBR Labor Market data from {FILE_PATH.name}")
    print()

    # Extract from report sheet
    df = extract_from_report_sheet()

    if not df.empty:
        # Save long format
        df.to_csv(OUTPUT_PATH, index=False)
        print(f"\nSaved long format to: {OUTPUT_PATH}")
        print(f"  Total records: {len(df)}")
        print(f"  Dimensions: {df.shape}")

        # Also save wide format
        df_wide = pivot_to_wide_format(df)
        wide_path = OUTPUT_PATH.with_name("kbr_labor_market_wide.csv")
        df_wide.to_csv(wide_path, index=False)
        print(f"\nSaved wide format to: {wide_path}")
        print(f"  Dimensions: {df_wide.shape}")

        # Generate summary
        generate_summary(df)

        return df
    else:
        print("\nERROR: No data extracted!")
        return pd.DataFrame()


if __name__ == "__main__":
    main()
