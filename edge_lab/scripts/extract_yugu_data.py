#!/usr/bin/env python3

import openpyxl
import csv
from pathlib import Path
from datetime import datetime
import warnings

warnings.filterwarnings("ignore")

EXCEL_FILE = "assets/charts/ОПР_статистика/Основная статистика ЮГУ.xlsx"
OUTPUT_FILE = "data/raw_yugu_dump.csv"


def extract_sheet_data(sheet_name, sheet):
    row1 = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    indicator = str(row1[0]) if row1[0] else "No indicator"

    date_row_idx = None
    for i in range(1, min(20, sheet.max_row + 1)):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if len(row) > 1 and row[1] and hasattr(row[1], "year"):
            date_row_idx = i
            break

    if not date_row_idx:
        return []

    kdk_row_idx = None
    for i in range(1, min(100, sheet.max_row + 1)):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if row[0] and str(row[0]).strip() == "Кдк":
            kdk_row_idx = i
            break

    if not kdk_row_idx:
        return []

    date_row = list(
        sheet.iter_rows(min_row=date_row_idx, max_row=date_row_idx, values_only=True)
    )[0]
    kdk_row = list(
        sheet.iter_rows(min_row=kdk_row_idx, max_row=kdk_row_idx, values_only=True)
    )[0]

    results = []
    for col_idx, (date_cell, val_cell) in enumerate(
        zip(date_row[1:], kdk_row[1:]), start=1
    ):
        if date_cell and val_cell is not None:
            date_str = (
                date_cell.strftime("%Y-%m-%d")
                if isinstance(date_cell, datetime)
                else str(date_cell)
            )
            results.append((date_str, float(val_cell), indicator, sheet_name))

    return results


def main():
    Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)

    total_records = 0
    sheets_processed = 0
    sheets_with_data = 0

    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Date", "Value", "Indicator", "Sheet"])

        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            sheets_processed += 1
            sheet = wb[sheet_name]

            if sheet.max_row < 15 or sheet.max_column < 5:
                continue

            data = extract_sheet_data(sheet_name, sheet)

            if data:
                sheets_with_data += 1
                writer.writerows(data)
                total_records += len(data)
                print(
                    f"Sheet {sheets_processed:3d}/{len(wb.sheetnames)}: {sheet_name:6s} -> {len(data):4d} records"
                )

        wb.close()

    print(f"\nSummary:")
    print(f"  Total sheets processed: {sheets_processed}")
    print(f"  Sheets with KBR data: {sheets_with_data}")
    print(f"  Total records extracted: {total_records}")
    print(f"  Output: {OUTPUT_FILE}")

    return sheets_with_data, total_records


if __name__ == "__main__":
    main()
