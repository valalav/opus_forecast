#!/usr/bin/env python3

import openpyxl
import csv
from pathlib import Path
from datetime import datetime
import warnings

warnings.filterwarnings("ignore")

EXCEL_FILE = "assets/charts/ОПР_статистика/Основная статистика ЮГУ.xlsx"
OUTPUT_FILE = "data/kbr_sectoral_details.csv"

REGION_CODES = {
    "РФ",
    "ЮМР",
    "ЮФО",
    "Ады",
    "Клм",
    "Крм",
    "Кдк",
    "Аст",
    "Влг",
    "Рос",
    "Сев",
    "СКФО",
    "Даг",
    "Инг",
    "Кбр",
    "Кчр",
    "Ост",
    "Чеч",
    "Ств",
}

METRIC_TYPES = {"г/г", "м/м", "БИ", "г/г AR", "м/м AR"}


def get_sheet_structure(sheet):
    for i in range(6, min(12, sheet.max_row + 1)):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if len(row) > 1 and row[0] and "Территория" in str(row[0]):
            if row[1] and "Кбр" in str(row[1]):
                return "direct"

    for i in range(1, min(50, sheet.max_row + 1)):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if row[0] and str(row[0]).strip() == "Кбр":
            return "block"

    return "unknown"


def extract_sheet_data_block(sheet_name, sheet):
    row1 = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    sheet_title = str(row1[0]) if row1[0] else sheet_name

    date_row_idx = None
    for i in range(1, min(15, sheet.max_row + 1)):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if len(row) > 1 and row[1] and hasattr(row[1], "year"):
            date_row_idx = i
            break

    if not date_row_idx:
        return [], "no_date_row"

    date_row = list(
        sheet.iter_rows(min_row=date_row_idx, max_row=date_row_idx, values_only=True)
    )[0]

    kbr_data_start_col = None
    for i in range(1, min(50, sheet.max_row + 1)):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if row[0] and str(row[0]).strip() == "Кбр":
            for j in range(1, min(15, len(row))):
                if row[j] is not None:
                    kbr_data_start_col = j
                    break
            break

    if not kbr_data_start_col:
        return [], "no_kbr_col"

    results = []
    current_indicator = sheet_title
    current_metric = None

    for row_idx in range(1, sheet.max_row + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[
            0
        ]

        if not row[0]:
            continue

        cell_val = str(row[0]).strip()

        if not cell_val:
            continue

        if cell_val == "Кбр":
            if current_indicator:
                for col_idx in range(kbr_data_start_col, len(row)):
                    date_cell = date_row[col_idx] if col_idx < len(date_row) else None
                    val_cell = row[col_idx]

                    if date_cell and val_cell is not None:
                        try:
                            date_str = (
                                date_cell.strftime("%Y-%m-%d")
                                if isinstance(date_cell, datetime)
                                else str(date_cell)
                            )
                            val = float(val_cell)
                            results.append(
                                (
                                    date_str,
                                    val,
                                    current_indicator,
                                    current_metric,
                                    sheet_name,
                                )
                            )
                        except (ValueError, TypeError):
                            continue

        elif cell_val in METRIC_TYPES:
            current_metric = cell_val

        elif ";" in cell_val and cell_val not in REGION_CODES:
            current_indicator = cell_val

    return results, "ok"


def extract_sheet_data_direct(sheet_name, sheet):
    row1 = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    sheet_title = str(row1[0]) if row1[0] else sheet_name

    date_row_idx = None
    for i in range(1, min(15, sheet.max_row + 1)):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if len(row) > 1 and row[1] and hasattr(row[1], "year"):
            date_row_idx = i
            break

    if not date_row_idx:
        return [], "no_date_row"

    date_row = list(
        sheet.iter_rows(min_row=date_row_idx, max_row=date_row_idx, values_only=True)
    )[0]

    kbr_data_start_col = None
    for i in range(date_row_idx + 1, sheet.max_row + 1):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        cell_val = str(row[0]).strip() if row[0] else ""
        if cell_val and ";" in cell_val:
            for j in range(1, min(10, len(row))):
                if row[j] is not None:
                    kbr_data_start_col = j
                    break
            break

    if not kbr_data_start_col:
        return [], "no_kbr_col"

    results = []
    current_indicator = sheet_title
    current_metric = None

    for row_idx in range(date_row_idx + 1, sheet.max_row + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[
            0
        ]

        if not row[0]:
            continue

        cell_val = str(row[0]).strip()

        if not cell_val:
            continue

        if cell_val in METRIC_TYPES:
            current_metric = cell_val
        elif cell_val not in REGION_CODES:
            if ";" in cell_val:
                current_indicator = cell_val

                for col_idx in range(kbr_data_start_col, len(row)):
                    date_cell = date_row[col_idx] if col_idx < len(date_row) else None
                    val_cell = row[col_idx]

                    if date_cell and val_cell is not None:
                        try:
                            date_str = (
                                date_cell.strftime("%Y-%m-%d")
                                if isinstance(date_cell, datetime)
                                else str(date_cell)
                            )
                            val = float(val_cell)
                            results.append(
                                (
                                    date_str,
                                    val,
                                    current_indicator,
                                    current_metric,
                                    sheet_name,
                                )
                            )
                        except (ValueError, TypeError):
                            continue

    return results, "ok"


def extract_sheet_data(sheet_name, sheet):
    structure = get_sheet_structure(sheet)

    if structure == "direct":
        return extract_sheet_data_direct(sheet_name, sheet)
    elif structure == "block":
        return extract_sheet_data_block(sheet_name, sheet)
    else:
        return [], "unknown_structure"


def main():
    Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)

    total_records = 0
    sheets_processed = 0
    sheets_with_data = 0

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    all_sheets = wb.sheetnames

    numeric_sheets = [name for name in all_sheets if name.isdigit()]

    numeric_ge_100 = [name for name in numeric_sheets if int(name) >= 100]
    wb.close()

    print(f"Total sheets in workbook: {len(all_sheets)}")
    print(f"Numeric sheets: {len(numeric_sheets)}")
    print(f"Numeric sheets >= 100: {len(numeric_ge_100)}")
    print(f"Processing {len(numeric_ge_100)} thematic sheets...\n")

    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Date", "Value", "Indicator", "Metric_Type", "Sheet"])

        batch_size = 10
        total_batches = (len(numeric_ge_100) + batch_size - 1) // batch_size

        for batch_idx in range(total_batches):
            start = batch_idx * batch_size
            end = min(start + batch_size, len(numeric_ge_100))
            batch = numeric_ge_100[start:end]

            wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

            for sheet_name in batch:
                sheets_processed += 1
                sheet = wb[sheet_name]

                if sheet.max_row < 10 or sheet.max_column < 3:
                    wb.close()
                    continue

                data, reason = extract_sheet_data(sheet_name, sheet)

                if data:
                    sheets_with_data += 1
                    writer.writerows(data)
                    total_records += len(data)
                    print(f"Sheet {sheet_name:3s}: {len(data):6d} records")

            wb.close()

    print(f"\nSummary:")
    print(f"  Total thematic sheets processed: {sheets_processed}")
    print(f"  Sheets with KBR data: {sheets_with_data}")
    print(f"  Total records extracted: {total_records}")
    print(f"  Output: {OUTPUT_FILE}")

    return sheets_with_data, total_records


if __name__ == "__main__":
    main()
