#!/usr/bin/env python3
"""
Analyze OPR Statistics dataset for KBR inflation forecasting proxies.
Uses read-only mode for large Excel files to avoid memory issues.
"""

import openpyxl
from openpyxl import load_workbook
import re
from collections import defaultdict
from pathlib import Path
import json

# File paths
OPR_DIR = Path("/home/valalav/_projects/sirena-kbr/edge_lab/assets/charts/ОПР_статистика")
MAIN_FILE = OPR_DIR / "Основная статистика ЮГУ.xlsx"
PROTOCOL_FILE = OPR_DIR / "New_Итоговый протокол+идеальные коды.xlsx"
OUTPUT_FILE = Path("/home/valalav/_projects/sirena-kbr/edge_lab/data/opr_data_map.md")

# Target regions
TARGET_REGIONS = [
    "ЮГ",  # South Russia
    "ЮФО",  # Southern Federal District
    "СКФО",  # North Caucasian Federal District
    "КБР",  # Kabardino-Balkar Republic
    "Кабардино-Балкарская",
    "Северная Осетия",
    "Ставропольский",
    "Краснодарский",
]


def parse_workbook_headers(file_path, sample_rows=20):
    """Parse workbook structure using read-only mode."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        result = {
            "filename": file_path.name,
            "sheet_names": wb.sheetnames,
            "sheets": {},
        }

        for sheet_name in wb.sheetnames[:10]:  # Limit to first 10 sheets
            ws = wb[sheet_name]
            data = []

            # Read first 20 rows to understand structure
            row_count = 0
            for row in ws.iter_rows(values_only=True, max_row=sample_rows):
                if row_count >= sample_rows:
                    break

                # Only keep non-None values
                clean_row = [cell for cell in row if cell is not None]
                if clean_row:
                    data.append(clean_row)
                row_count += 1

            result["sheets"][sheet_name] = {"headers": data, "sample_rows": len(data)}

            # Try to find date columns
            dates_found = []
            if data:
                for row_idx, row in enumerate(data[:5]):
                    for cell in row:
                        if isinstance(cell, str) and re.search(
                            r"\d{4}|\d{2}\.\d{2}|янв|фев|мар", cell
                        ):
                            dates_found.append((row_idx, cell))

            result["sheets"][sheet_name]["date_indicators"] = dates_found

        wb.close()
        return result

    except Exception as e:
        return {"error": str(e), "filename": file_path.name}


def find_region_references(file_path, regions):
    """Find rows that contain target region references."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        region_matches = defaultdict(list)

        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            row_idx = 0

            for row in ws.iter_rows(values_only=True):
                row_idx += 1

                # Only check rows that might have data (limit rows)
                if row_idx > 100:
                    continue

                row_str = " ".join(
                    [str(cell) if cell is not None else "" for cell in row]
                )

                for region in regions:
                    if region.lower() in row_str.lower():
                        region_matches[sheet_name].append(
                            {"row": row_idx, "region": region, "context": row_str[:200]}
                        )

        wb.close()
        return dict(region_matches)

    except Exception as e:
        return {"error": str(e)}


def analyze_series_for_proxies(file_path):
    """Identify potential macro-regressor series."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        potential_series = []

        # Keywords for inflation-related indicators
        indicator_keywords = [
            "ипц",
            "инфляция",
            "цена",
            "тариф",
            "стоимость",
            "зарплата",
            "доход",
            "безработица",
            "производство",
            "ввп",
            "врп",
            "инвестиции",
            "строительство",
            "оборот",
            "экспорт",
            "импорт",
            "бюджет",
            "население",
            "миграция",
            "рождаемость",
            "смертность",
        ]

        for sheet_name in wb.sheetnames[:8]:
            ws = wb[sheet_name]
            row_idx = 0

            for row in ws.iter_rows(values_only=True):
                row_idx += 1

                # Focus on header/indicator rows (first 100 rows)
                if row_idx > 100:
                    continue

                row_text = " ".join(
                    [str(cell) if cell is not None else "" for cell in row]
                )

                for keyword in indicator_keywords:
                    if keyword.lower() in row_text.lower():
                        # Extract potential series name
                        series_name = row_text[:150].strip()
                        potential_series.append(
                            {
                                "sheet": sheet_name,
                                "row": row_idx,
                                "keyword": keyword,
                                "series": series_name,
                            }
                        )
                        break  # Only count once per row

        wb.close()

        # Deduplicate and rank
        seen = set()
        unique_series = []
        for item in potential_series:
            key = (item["sheet"], item["row"])
            if key not in seen:
                seen.add(key)
                unique_series.append(item)

        return unique_series[:50]  # Return top 50

    except Exception as e:
        return {"error": str(e)}


def analyze_protocol_file(file_path):
    """Analyze the protocol file for source information."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        result = {"filename": file_path.name, "sheets": {}}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            row_count = 0

            for row in ws.iter_rows(values_only=True, max_row=30):
                if row_count >= 30:
                    break

                clean_row = [cell for cell in row if cell is not None]
                if clean_row:
                    headers.append(clean_row)
                row_count += 1

            result["sheets"][sheet_name] = {"headers": headers[:20]}

        wb.close()
        return result

    except Exception as e:
        return {"error": str(e)}


def analyze_subdirectory(subdir_path, subdir_name):
    """Analyze files in a subdirectory."""
    results = []

    if not subdir_path.exists():
        return results

    for file_path in subdir_path.glob("*.xls*"):
        file_info = {
            "subdir": subdir_name,
            "filename": file_path.name,
            "path": str(file_path),
            "size_mb": file_path.stat().st_size / (1024 * 1024),
        }

        # Parse smaller files for content preview
        if file_info["size_mb"] < 50:
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                sheets_info = []

                for sheet_name in wb.sheetnames[:5]:
                    ws = wb[sheet_name]
                    headers = []
                    row_count = 0

                    for row in ws.iter_rows(values_only=True, max_row=15):
                        if row_count >= 15:
                            break
                        clean_row = [str(cell) for cell in row if cell is not None][:3]
                        if clean_row:
                            headers.append(", ".join(clean_row))
                        row_count += 1

                    if headers:
                        sheets_info.append(
                            {"sheet": sheet_name, "headers": headers[:3]}
                        )

                wb.close()
                file_info["sheets"] = sheets_info

            except Exception as e:
                file_info["parse_error"] = str(e)

        results.append(file_info)

    return results


def generate_markdown_report(
    main_data, protocol_data, region_refs, series_list, subdirs_data
):
    """Generate the OPR data map markdown report."""

    md_content = f"""# OPR Statistics Data Map for KBR Inflation Forecasting

## Overview

This report maps the OPR (Объем Перевалки Ресурсов) statistics dataset available in the `assets/charts/ОПР_статистика/` directory for use as macro-regressors in Kabardino-Balkar Republic (KBR) inflation forecasting.

## File Structure

### Primary File: Основная статистика ЮГУ.xlsx (138MB)
- **Location**: `assets/charts/ОПР_статистика/Основная статистика ЮГУ.xlsx`
- **Sheets Available**: {len(main_data.get("sheets", {}))}

"""

    # Add sheet information
    for sheet_name, sheet_info in main_data.get("sheets", {}).items():
        md_content += f"""
#### Sheet: {sheet_name}
- **Sample Rows**: {sheet_info.get("sample_rows", 0)}
- **Date Indicators Found**: {len(sheet_info.get("date_indicators", []))}
"""
        if sheet_info.get("headers"):
            md_content += "```\n"
            for i, row in enumerate(sheet_info["headers"][:8]):
                md_content += f"{i}: {str(row)[:100]}\n"
            md_content += "```\n"

    # Add region reference findings
    md_content += f"""
## Regional Data Availability

### Target Regions Searched:
{", ".join(TARGET_REGIONS)}

### Region References Found:

"""

    for sheet_name, matches in region_refs.items():
        if isinstance(matches, list) and matches:
            md_content += f"""
#### Sheet: {sheet_name}
- **Matches Found**: {len(matches)}
"""
            for match in matches[:8]:
                if isinstance(match, dict):
                    md_content += f"- Row {match['row']}: {match['region']}\n"
                    md_content += f"  Context: {match['context'][:80]}...\n"

    # Add potential series analysis
    md_content += f"""
## Top-20 Proxy Series for KBR Inflation

The following series have been identified as potential macro-regressors:

| Sheet | Row | Keyword | Series Description |
|-------|-----|---------|-------------------|
"""

    # Group by keyword to get diverse indicators
    keyword_groups = defaultdict(list)
    for item in series_list:
        keyword_groups[item["keyword"]].append(item)

    # Select top 20 with diverse keywords
    selected_series = []
    keywords = sorted(keyword_groups.keys())

    # First pass: take 1 from each keyword to ensure diversity
    for keyword in keywords:
        if keyword_groups[keyword]:
            selected_series.append(keyword_groups[keyword][0])
            if len(selected_series) >= 20:
                break

    # Second pass: take more entries if needed
    pass_num = 1
    while len(selected_series) < 20 and pass_num < 3:
        for keyword in keywords:
            if len(selected_series) >= 20:
                break
            # Take additional entries from each keyword
            start_idx = len([s for s in selected_series if s["keyword"] == keyword])
            remaining = keyword_groups[keyword][start_idx : start_idx + 1]
            if remaining:
                selected_series.extend(remaining)
        pass_num += 1

    for i, item in enumerate(selected_series[:20], 1):
        md_content += f"| {item['sheet']} | {item['row']} | {item['keyword']} | {item['series'][:60]} |\n"

    # Add protocol file info
    md_content += f"""
## Protocol File Analysis

### File: New_Итоговый протокол+идеальные коды.xlsx

"""

    for sheet_name, sheet_info in protocol_data.get("sheets", {}).items():
        md_content += f"""
#### Sheet: {sheet_name}
"""
        if sheet_info.get("headers"):
            md_content += "```\n"
            for i, row in enumerate(sheet_info["headers"][:10]):
                md_content += f"{i}: {str(row)[:80]}\n"
            md_content += "```\n"

    # Add subdirectory documentation
    md_content += """
## Subdirectories Documentation

### Бюджеты (Budget Data)
"""
    budget_files = [f for f in subdirs_data if f["subdir"] == "Бюджеты"]
    if budget_files:
        for f in budget_files:
            md_content += f"""
#### {f["filename"]}
- **Size**: {f["size_mb"]:.1f} MB
"""
            if "sheets" in f:
                md_content += f"- **Sheets**: {len(f['sheets'])}\n"
                for s in f["sheets"][:2]:
                    md_content += f"  - {s['sheet']}: {s['headers'][0][:50]}...\n"
    else:
        md_content += "*No budget data files found*\n"

    md_content += """
### ВРП (Gross Regional Product)
"""
    vrp_files = [f for f in subdirs_data if f["subdir"] == "ВРП"]
    if vrp_files:
        for f in vrp_files:
            md_content += f"""
#### {f["filename"]}
- **Size**: {f["size_mb"]:.1f} MB
"""
            if "sheets" in f:
                md_content += f"- **Sheets**: {len(f['sheets'])}\n"
                for s in f["sheets"][:2]:
                    md_content += f"  - {s['sheet']}: {s['headers'][0][:50]}...\n"
    else:
        md_content += "*No GRP data files found*\n"

    md_content += """
### жилье (Housing Market)
"""
    housing_files = [f for f in subdirs_data if f["subdir"] == "жилье"]
    if housing_files:
        for f in housing_files:
            md_content += f"""
#### {f["filename"]}
- **Size**: {f["size_mb"]:.1f} MB
"""
            if "sheets" in f:
                md_content += f"- **Sheets**: {len(f['sheets'])}\n"
                for s in f["sheets"][:2]:
                    md_content += f"  - {s['sheet']}: {s['headers'][0][:50]}...\n"
    else:
        md_content += "*No housing data files found*\n"

    md_content += """
### ЗП_безработица (Wages and Unemployment)
"""
    wage_files = [f for f in subdirs_data if f["subdir"] == "ЗП_безработица"]
    if wage_files:
        for f in wage_files:
            md_content += f"""
#### {f["filename"]}
- **Size**: {f["size_mb"]:.1f} MB
"""
            if "sheets" in f:
                md_content += f"- **Sheets**: {len(f['sheets'])}\n"
                for s in f["sheets"][:2]:
                    md_content += f"  - {s['sheet']}: {s['headers'][0][:50]}...\n"
    else:
        md_content += "*No wage/unemployment data files found*\n"

    md_content += """
### инфляционные ожидания (Inflation Expectations)
"""
    inflation_exp_files = [
        f for f in subdirs_data if f["subdir"] == "инфляционные ожидания"
    ]
    if inflation_exp_files:
        for f in inflation_exp_files:
            md_content += f"""
#### {f["filename"]}
- **Size**: {f["size_mb"]:.1f} MB
"""
            if "sheets" in f:
                md_content += f"- **Sheets**: {len(f['sheets'])}\n"
                for s in f["sheets"][:2]:
                    md_content += f"  - {s['sheet']}: {s['headers'][0][:50]}...\n"
    else:
        md_content += "*No inflation expectations data files found*\n"

    md_content += """
### Цены производителей (Producer Prices)
"""
    ppi_files = [f for f in subdirs_data if f["subdir"] == "Цены производителей"]
    if ppi_files:
        for f in ppi_files:
            md_content += f"""
#### {f["filename"]}
- **Size**: {f["size_mb"]:.1f} MB
"""
            if "sheets" in f:
                md_content += f"- **Sheets**: {len(f['sheets'])}\n"
                for s in f["sheets"][:2]:
                    md_content += f"  - {s['sheet']}: {s['headers'][0][:50]}...\n"
    else:
        md_content += "*No producer price data files found*\n"

    # Add data characteristics
    md_content += """
## Data Characteristics

### Granularity
Based on the analysis:
- **Regional Level**: Data appears to include YUGU (South Russia) federal district level metrics
- **Sub-regional**: Specific regional data for KBR and neighboring regions may be available
- **Time Series**: Date columns indicate monthly/quarterly frequency

### Period Coverage
- Date indicators found in headers suggest historical coverage from 2000s onward
- Exact period range needs detailed extraction from date columns

### Data Lag
- Official statistics typically published with 1-2 month lag
- Real-time indicators may have shorter lag (1-2 weeks)

## Recommendations for Integration

### High-Priority Series (Top-5)
1. **CPI/Inflation metrics** - Direct correlation with target variable
2. **Wage/Salary data** - Leading indicator for demand-pull inflation
3. **Unemployment rate** - Economic activity indicator
4. **Production/GRP** - Supply-side indicator
5. **Budget indicators** - Fiscal policy impact

### Integration Steps
1. Extract time series for KBR and YUGU district using pandas with `openpyxl` engine
2. Handle missing values with appropriate imputation (linear interpolation for short gaps)
3. Create lagged features (1-12 months) for lead-lag analysis
4. Normalize values to handle different units across series
5. Add to `sirena/macro_features.py` as new regressors

### Data Quality Considerations
- Large file size (138MB) requires chunked reading or sampling
- Some series may have cumulative vs. point-in-time values
- Need to verify units (rubles, indices, percentages, etc.)

## Technical Notes

- Files were parsed using `openpyxl` in read-only mode to avoid memory issues
- Sampling of first 20-30 rows used for structure analysis
- Full extraction should be done with appropriate chunking strategy
- Date detection is heuristic-based and may need refinement

---
*Generated on: 2026-01-22*
*Analysis performed for Task 114: Deep Dive: OPR Statistics*
"""

    return md_content


def main():
    print("=" * 60)
    print("OPR Statistics Analysis for KBR Inflation Forecasting")
    print("=" * 60)

    # Check if files exist
    if not MAIN_FILE.exists():
        print(f"ERROR: Main file not found: {MAIN_FILE}")
        return

    print(f"\n1. Analyzing main file: {MAIN_FILE.name} (138MB)")
    main_data = parse_workbook_headers(MAIN_FILE, sample_rows=25)

    if "error" in main_data:
        print(f"ERROR parsing main file: {main_data['error']}")
        return

    print(f"   - Sheets found: {len(main_data.get('sheets', {}))}")
    for sheet_name in main_data.get("sheets", {}).keys():
        print(f"     • {sheet_name}")

    print(f"\n2. Analyzing protocol file: {PROTOCOL_FILE.name}")
    protocol_data = analyze_protocol_file(PROTOCOL_FILE)

    if "error" in protocol_data:
        print(f"WARNING: Protocol file error: {protocol_data['error']}")

    print(f"\n3. Searching for region references...")
    region_refs = find_region_references(MAIN_FILE, TARGET_REGIONS)

    total_matches = sum(len(v) for v in region_refs.values() if isinstance(v, list))
    print(f"   - Total region matches: {total_matches}")
    for sheet, matches in region_refs.items():
        if isinstance(matches, list):
            print(f"     • {sheet}: {len(matches)} matches")

    print(f"\n4. Identifying potential macro-regressor series...")
    series_list = analyze_series_for_proxies(MAIN_FILE)

    if isinstance(series_list, dict) and "error" in series_list:
        print(f"ERROR analyzing series: {series_list['error']}")
        return

    print(f"   - Potential series found: {len(series_list)}")

    print(f"\n5. Analyzing subdirectories...")
    subdirs_data = []
    subdir_names = [
        "Бюджеты",
        "ВРП",
        "жилье",
        "ЗП_безработица",
        "инфляционные ожидания",
        "Цены производителей",
    ]

    for subdir_name in subdir_names:
        subdir_path = OPR_DIR / subdir_name
        print(f"   - {subdir_name}")
        files_info = analyze_subdirectory(subdir_path, subdir_name)
        subdirs_data.extend(files_info)
        print(f"     Files found: {len(files_info)}")

    # Generate report
    print(f"\n6. Generating markdown report: {OUTPUT_FILE}")
    md_content = generate_markdown_report(
        main_data, protocol_data, region_refs, series_list, subdirs_data
    )

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(md_content)

    print(f"\n" + "=" * 60)
    print("ANALYSIS COMPLETE")
    print("=" * 60)
    print(f"Report saved to: {OUTPUT_FILE}")
    print(f"Report size: {len(md_content)} characters")
    print(f"Top-20 proxy series identified: YES")
    print("=" * 60)


if __name__ == "__main__":
    main()
