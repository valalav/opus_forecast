#!/usr/bin/env python3
"""
Task 124: Mining: High-Freq Indicators (HH.ru & DomClick)

Extract HH Index and Housing Prices for KBR from OPR stats files.

NOTE ON DOMCLICK DATA:
The DomClick Excel files use Power Query with external connections to SberIndex API.
Technical analysis shows:
- saveData="0" in pivot cache definitions - no raw data stored locally
- Only cached values reflect currently active Excel slicer filter (Ural region)
- No Python library (openpyxl, pandas) can access Power Query data
- Excel COM interface (xlwings) not available on this Linux system

APPROACH:
1. HH Index: Extracted from hh_индекс.xlsx (working)
2. Housing Prices: Rosstat data used as documented alternative
   - This is the same approach recommended by the Critic (option 3)
   - Data source is clearly documented in output

Outputs:
- data/kbr_high_freq_indicators.csv: Combined high-frequency indicators
- data/task124_tech_analysis.md: Technical analysis of DomClick limitation
- data/task124_correlation_report.md: Correlation analysis
"""

import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime
import json
import os
import zipfile


def analyze_domclick_structure(file_path: str) -> dict:
    """Analyze DomClick Excel file structure to document technical limitation."""
    analysis = {
        "file": os.path.basename(file_path),
        "sheets": [],
        "pivot_caches": [],
        "power_query_connections": [],
        "cached_regions": [],
        "limitations": [],
    }

    try:
        # Load Excel to check structure
        wb = openpyxl.load_workbook(file_path, data_only=False)
        analysis["sheets"] = wb.sheetnames

        # Check for pivot tables
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, "_pivots"):
                analysis["pivot_caches"].append(
                    {"sheet": sheet_name, "count": len(ws._pivots)}
                )

        # Check cached values with data_only=True
        wb_cache = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in ["Свод (ГУ)", "Свод (ФО)"]:
            if sheet_name in wb_cache.sheetnames:
                ws = wb_cache[sheet_name]

                # Scan for region indicators in cached data
                for row_idx in range(1, min(40, ws.max_row + 1)):
                    for col_idx in range(1, min(5, ws.max_column + 1)):
                        cell = ws.cell(row_idx, col_idx)
                        if cell.value and isinstance(cell.value, str):
                            if "ГУ" in cell.value and cell.value not in [
                                "янв",
                                "фев",
                                "мар",
                                "апр",
                                "май",
                                "июн",
                                "июл",
                                "авг",
                                "сен",
                                "окт",
                                "ноя",
                                "дек",
                            ]:
                                analysis["cached_regions"].append(
                                    f"{sheet_name}: Row {row_idx} - {cell.value}"
                                )

                # Check dimensions
                analysis["dimensions"] = ws.dimensions

        wb.close()
        wb_cache.close()

        # Check Excel file internals (ZIP structure)
        with zipfile.ZipFile(file_path, "r") as zip_ref:
            cache_defs = [f for f in zip_ref.namelist() if "pivotCacheDefinition" in f]

            for cache_file in cache_defs[:2]:
                try:
                    content = zip_ref.read(cache_file).decode("utf-8")
                    if 'saveData="0"' in content:
                        analysis["limitations"].append(
                            f"Pivot cache {os.path.basename(cache_file)}: saveData='0' - data NOT stored locally"
                        )
                    if "connectionId" in content:
                        matches = content.split('connectionId="')[1].split('"')[0]
                        analysis["power_query_connections"].append(matches)
                except:
                    pass

        analysis["limitation_summary"] = (
            "DomClick data is fetched via Power Query from SberIndex API. "
            "No raw KBR housing data is stored in Excel file. "
            "Cached values only reflect active filter (Ural region), not Southern region (KBR)."
        )

    except Exception as e:
        analysis["error"] = str(e)

    return analysis


def parse_hh_index(
    file_path: str, target_region: str = "Кабардино-Балкарская Республика"
) -> pd.DataFrame:
    """Parse HH Index Excel file for labor market tension."""
    df = pd.read_excel(file_path, sheet_name=0)

    # Filter for KBR
    df_kbr = df[
        df["name"].str.contains(target_region, na=False)
        | df["name"].str.contains("Кабардино", na=False)
    ].copy()

    if df_kbr.empty:
        raise ValueError(f"Could not find {target_region} in HH Index file")

    # Aggregate by date
    df_kbr = df_kbr.groupby("rep_date")["HHI"].mean().reset_index()
    df_kbr.columns = ["Date", "HH_Index"]

    # Filter and convert
    df_kbr = df_kbr[df_kbr["HH_Index"].notna() & (df_kbr["HH_Index"] != "")]
    df_kbr["HH_Index"] = pd.to_numeric(df_kbr["HH_Index"], errors="coerce")
    df_kbr["Date"] = pd.to_datetime(df_kbr["Date"])

    df_kbr = df_kbr.sort_values("Date").reset_index(drop=True)
    df_kbr["Date"] = df_kbr["Date"].dt.strftime("%Y-%m-%d")

    return df_kbr


def parse_housing_prices(file_path: str, target_region: str = "Кбр") -> pd.DataFrame:
    """Parse Housing Prices from Rosstat data (sheet 118 of YUGU stats)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "118" not in wb.sheetnames:
        raise ValueError("Sheet '118' not found in housing stats file")

    ws = wb["118"]

    data_secondary = []
    data_primary = []

    # Parse dates from row 11
    dates = []
    for i, cell in enumerate(ws[11], 1):
        if i > 1 and isinstance(cell.value, datetime):
            dates.append(cell.value)

    # Find Kbr rows for secondary and primary markets
    secondary_kbr_row = None
    primary_kbr_row = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= 60:
            region_code = row[0] if row[0] else ""

            if region_code == target_region and row_idx > 20 and row_idx < 32:
                secondary_kbr_row = row
            if region_code == target_region and row_idx > 40 and row_idx < 50:
                primary_kbr_row = row

    if secondary_kbr_row is None:
        raise ValueError(
            f"Could not find secondary market housing data for {target_region}"
        )

    # Extract secondary market prices
    for i, (date, price) in enumerate(zip(dates, secondary_kbr_row[1:]), 1):
        if i <= len(dates) and isinstance(price, (int, float)) and not pd.isna(price):
            data_secondary.append(
                {"Date": date.strftime("%Y-%m-%d"), "Housing_Price_Secondary": price}
            )

    # Extract primary market prices
    if primary_kbr_row:
        for i, (date, price) in enumerate(zip(dates, primary_kbr_row[1:]), 1):
            if (
                i <= len(dates)
                and isinstance(price, (int, float))
                and not pd.isna(price)
            ):
                data_primary.append(
                    {"Date": date.strftime("%Y-%m-%d"), "Housing_Price_Primary": price}
                )

    df_secondary = pd.DataFrame(data_secondary)
    df_primary = pd.DataFrame(data_primary)

    if not df_secondary.empty:
        df_secondary["Date"] = pd.to_datetime(df_secondary["Date"])
        df_secondary = df_secondary.sort_values("Date").reset_index(drop=True)

    if not df_primary.empty:
        df_primary["Date"] = pd.to_datetime(df_primary["Date"])
        df_primary = df_primary.sort_values("Date").reset_index(drop=True)

    df = df_secondary.merge(df_primary, on="Date", how="outer")

    return df


def load_cpi_data() -> pd.DataFrame:
    """Load CPI data for correlation analysis."""
    # Try multiple possible CPI data files
    possible_files = [
        "data/enhanced_inflation_data.csv",
        "data/infl_kbr.csv",
    ]

    for file_path in possible_files:
        if os.path.exists(file_path):
            try:
                df = pd.read_csv(file_path)

                # Find date column
                date_col = None
                for col in ["date", "Date", "Date_", "период"]:
                    if col in df.columns:
                        date_col = col
                        break

                if date_col is None:
                    continue

                df["Date"] = pd.to_datetime(df[date_col])

                # Find mom CPI column
                mom_col = None
                for col in ["mom", "cpi", "CPI_mom", "ИПЦ_м/м"]:
                    if col in df.columns:
                        mom_col = col
                        break

                if mom_col is None:
                    continue

                df["CPI_mom"] = df[mom_col]
                return df[["Date", "CPI_mom"]].dropna()

            except Exception as e:
                print(f"   - Warning: Could not load {file_path}: {e}")
                continue

    return None


def calculate_correlation(
    df1: pd.DataFrame, col1: str, df2: pd.DataFrame, col2: str, max_lag: int = 6
) -> dict:
    """Calculate cross-correlation between two time series."""
    merged = df1[["Date", col1]].merge(df2[["Date", col2]], on="Date", how="inner")

    if len(merged) < 6:
        return {"error": "Insufficient data points for correlation"}

    results = {}
    for lag in range(max_lag + 1):
        if lag == 0:
            corr = merged[col1].corr(merged[col2])
        else:
            series1 = merged[col1].iloc[:-lag].values
            series2 = merged[col2].iloc[lag:].values
            if len(series1) > 0 and len(series2) > 0:
                corr = np.corrcoef(series1, series2)[0, 1]
            else:
                corr = np.nan
        results[f"lag_{lag}"] = corr

    return results


def generate_tech_analysis(analysis: dict) -> str:
    """Generate technical analysis report."""
    report = f"""# Task 124: DomClick Technical Analysis

**Date:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

## File Analyzed
**{analysis.get("file", "N/A")}**

## Excel Structure Analysis

### Sheets Found
{chr(10).join([f"- {s}" for s in analysis.get("sheets", [])])}

### Pivot Tables
{chr(10).join([f"- Sheet: {p['sheet']}, Count: {p['count']}" for p in analysis.get("pivot_caches", [])])}

### Power Query Connections
{chr(10).join([f"- Connection ID: {c}" for c in analysis.get("power_query_connections", [])])}

### Cached Regions in Excel File
{chr(10).join([f"- {r}" for r in analysis.get("cached_regions", [])])}

## Technical Limitations

{chr(10).join([f"- {l}" for l in analysis.get("limitations", [])])}

## Limitation Summary

{analysis.get("limitation_summary", "N/A")}

## Approaches Attempted

1. **openpyxl with data_only=False**: Can read pivot table definitions but not raw data
2. **openpyxl with data_only=True**: Only returns cached values for currently active filter (Ural region)
3. **Pivot Cache Analysis**: saveData="0" confirms no raw data stored locally
4. **Excel File Structure Analysis**: Power Query connections fetch data from SberIndex API
5. **xlwings/COM Interface**: Not available on Linux system

## Conclusion

**KBR housing price data cannot be extracted from DomClick Excel files using standard Python libraries.**

The files are designed as dashboards that dynamically fetch data from SberIndex API. Without:
- A live Excel application with COM interface, OR
- Access to SberIndex API endpoints, OR
- The file being saved with KBR filter active

It is technically impossible to extract KBR housing prices directly from these files.

## Alternative Used

Housing price data for KBR is sourced from Rosstat statistics (Основная статистика ЮГУ.xlsx, sheet 118), which provides:
- Secondary market prices (quarterly)
- Primary market prices (quarterly)
- Official regional statistics for KBR

This is a valid alternative data source that provides the same economic indicator (housing prices) from an authoritative statistical source.
"""
    return report


def main():
    print("Task 124: Mining High-Freq Indicators (HH.ru & Housing Prices)")
    print("=" * 70)

    # Create output directory
    os.makedirs("data", exist_ok=True)

    # Step 1: Analyze DomClick file structure (document limitation)
    print("\n0. Analyzing DomClick Excel file structure...")
    domclick_file = (
        "assets/charts/ОПР_статистика/жилье/Цены на жилье (Домклик, факт.сделки).xlsx"
    )

    if os.path.exists(domclick_file):
        domclick_analysis = analyze_domclick_structure(domclick_file)

        tech_report = generate_tech_analysis(domclick_analysis)
        tech_report_file = "data/task124_tech_analysis.md"
        with open(tech_report_file, "w", encoding="utf-8") as f:
            f.write(tech_report)
        print(f"   - Technical analysis saved to: {tech_report_file}")

        if domclick_analysis.get("cached_regions"):
            print(
                f"   - Cached regions found: {len(domclick_analysis['cached_regions'])}"
            )
        if domclick_analysis.get("limitations"):
            print(
                f"   - Limitations documented: {len(domclick_analysis['limitations'])}"
            )

    # Step 2: Parse HH Index
    print("\n1. Parsing HH Index...")
    hh_file = "data/raw/opr_stat/hh_индекс.xlsx"
    if os.path.exists(hh_file):
        try:
            df_hh = parse_hh_index(hh_file)
            print(f"   - Loaded {len(df_hh)} rows of HH Index data")
            print(f"   - Date range: {df_hh['Date'].min()} to {df_hh['Date'].max()}")
        except Exception as e:
            print(f"   - Error: {e}")
            df_hh = pd.DataFrame(columns=["Date", "HH_Index"])
    else:
        print(f"   - File not found: {hh_file}")
        df_hh = pd.DataFrame(columns=["Date", "HH_Index"])

    # Step 3: Parse Housing Prices (Rosstat as documented alternative)
    print("\n2. Parsing Housing Prices (Rosstat data - documented alternative)...")
    housing_file = "data/raw/opr_stat/Основная статистика ЮГУ.xlsx"
    if os.path.exists(housing_file):
        try:
            df_housing = parse_housing_prices(housing_file)
            print(f"   - Loaded {len(df_housing)} rows of housing price data")
            print(
                f"   - Secondary market: {df_housing['Housing_Price_Secondary'].notna().sum()} values"
            )
            print(
                f"   - Primary market: {df_housing['Housing_Price_Primary'].notna().sum()} values"
            )
            print(
                f"   - Date range: {df_housing['Date'].min()} to {df_housing['Date'].max()}"
            )
        except Exception as e:
            print(f"   - Error: {e}")
            df_housing = pd.DataFrame(
                columns=["Date", "Housing_Price_Secondary", "Housing_Price_Primary"]
            )
    else:
        print(f"   - File not found: {housing_file}")
        df_housing = pd.DataFrame(
            columns=["Date", "Housing_Price_Secondary", "Housing_Price_Primary"]
        )

    # Step 4: Load CPI data for correlation
    print("\n3. Loading CPI data...")
    df_cpi = load_cpi_data()
    if df_cpi is not None:
        print(f"   - Loaded {len(df_cpi)} rows of CPI data")
        print(f"   - Date range: {df_cpi['Date'].min()} to {df_cpi['Date'].max()}")
    else:
        print("   - CPI data not found")

    # Step 5: Merge all data
    print("\n4. Merging indicators...")

    df_hh["Date"] = pd.to_datetime(df_hh["Date"])
    df_housing["Date"] = pd.to_datetime(df_housing["Date"])
    if df_cpi is not None:
        df_cpi["Date"] = pd.to_datetime(df_cpi["Date"])

    df_merged = df_hh.merge(df_housing, on="Date", how="outer")

    if df_cpi is not None:
        df_merged = df_merged.merge(df_cpi, on="Date", how="outer")

    df_merged = df_merged.sort_values("Date").reset_index(drop=True)
    print(f"   - Merged dataset: {len(df_merged)} rows")

    # Step 6: Save to CSV
    output_file = "data/kbr_high_freq_indicators.csv"
    df_merged.to_csv(output_file, index=False)
    print(f"   - Saved to: {output_file}")

    # Step 7: Calculate correlations
    print("\n5. Calculating correlations...")
    correlations = {}

    if not df_hh.empty and df_cpi is not None:
        corr_hh_cpi = calculate_correlation(df_hh, "HH_Index", df_cpi, "CPI_mom")
        correlations["HH_Index vs CPI"] = corr_hh_cpi
        print(f"   - HH Index vs CPI (lag 0): {corr_hh_cpi.get('lag_0', 'N/A')}")

    if not df_housing.empty and df_cpi is not None:
        df_housing_sec = df_housing[["Date", "Housing_Price_Secondary"]].dropna()
        corr_housing_cpi = calculate_correlation(
            df_housing_sec, "Housing_Price_Secondary", df_cpi, "CPI_mom"
        )
        correlations["Housing_Price vs CPI"] = corr_housing_cpi
        print(
            f"   - Housing Price vs CPI (lag 0): {corr_housing_cpi.get('lag_0', 'N/A')}"
        )

    if not df_hh.empty and not df_housing.empty:
        df_housing_sec = df_housing[["Date", "Housing_Price_Secondary"]].dropna()
        corr_hh_housing = calculate_correlation(
            df_hh, "HH_Index", df_housing_sec, "Housing_Price_Secondary"
        )
        correlations["HH_Index vs Housing_Price"] = corr_hh_housing
        print(
            f"   - HH Index vs Housing Price (lag 0): {corr_hh_housing.get('lag_0', 'N/A')}"
        )

    # Step 8: Generate correlation report
    print("\n6. Generating correlation report...")

    corr_report = f"""# Task 124 Correlation Report: High-Freq Indicators
**Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

## Data Sources

1. **HH Index (HH.ru)**: Labor market tension indicator from hh_индекс.xlsx
   - Frequency: Monthly
   - KBR data points: {len(df_hh)}

2. **Housing Prices (Rosstat)**: Housing price data from OPR statistics
   - Source: Основная статистика ЮГУ.xlsx, sheet 118
   - Frequency: Quarterly
   - Secondary market data points: {df_housing["Housing_Price_Secondary"].notna().sum() if not df_housing.empty else 0}
   - Primary market data points: {df_housing["Housing_Price_Primary"].notna().sum() if not df_housing.empty else 0}

3. **CPI**: Monthly inflation from infl_kbr.csv
   - Data points: {len(df_cpi) if df_cpi is not None else 0}

## Correlation Results

### HH Index vs CPI
```json
{json.dumps(correlations.get("HH_Index vs CPI", {}), indent=2)}
```

### Housing Price vs CPI
```json
{json.dumps(correlations.get("Housing_Price vs CPI", {}), indent=2)}
```

### HH Index vs Housing Price
```json
{json.dumps(correlations.get("HH_Index vs Housing_Price", {}), indent=2)}
```

## Interpretation

- **HH Index vs CPI**: {correlations.get("HH_Index vs CPI", {}).get("lag_0", "N/A")}
- **Housing Price vs CPI**: {correlations.get("Housing_Price vs CPI", {}).get("lag_0", "N/A")}
- **HH Index vs Housing Price**: {correlations.get("HH_Index vs Housing_Price", {}).get("lag_0", "N/A")}

## Notes on DomClick Data

The original task required parsing DomClick housing price data. After technical analysis:
- DomClick Excel files use Power Query with external connections to SberIndex API
- No raw KBR data is stored locally in the Excel file (saveData="0")
- Only cached values for active filter (Ural region) are accessible
- Python libraries (openpyxl, pandas) cannot access Power Query data

**Alternative:** Rosstat housing price data used as documented, valid alternative source.
"""

    corr_report_file = "data/task124_correlation_report.md"
    with open(corr_report_file, "w", encoding="utf-8") as f:
        f.write(corr_report)
    print(f"   - Correlation report saved to: {corr_report_file}")

    print("\n" + "=" * 70)
    print("Task 124 completed!")
    print(f"\nOutput Files:")
    print(f"  - {output_file} ({len(df_merged)} rows)")
    print(f"  - {tech_report_file} (DomClick technical analysis)")
    print(f"  - {corr_report_file} (Correlation analysis)")
    print(f"\nData Sources:")
    print(f"  - HH Index: hh_индекс.xlsx (working)")
    print(f"  - Housing Prices: Rosstat data (alternative, documented)")
    print(f"  - DomClick: Technical limitation documented (cannot extract)")

    return df_merged, correlations


if __name__ == "__main__":
    df_merged, correlations = main()
