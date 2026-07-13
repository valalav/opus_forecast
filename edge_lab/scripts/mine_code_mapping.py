#!/usr/bin/env python3
"""
Mining: Code Mapping & Protocol (Task 125)
Process 'New_Итоговый протокол+идеальные коды.xlsx' to extract indicator-to-code mappings.
"""

import pandas as pd
import json
from pathlib import Path
from openpyxl import load_workbook

# Configuration
INPUT_FILE = Path("data/raw/opr_stat/New_Итоговый протокол+идеальные коды.xlsx")
OUTPUT_FILE = Path("data/indicator_mapping_registry.csv")
METADATA_FILE = Path("data/mapping_metadata.json")


def analyze_excel_structure():
    """Analyze the Excel file structure to understand the data layout."""
    print(f"Analyzing: {INPUT_FILE}")

    wb = load_workbook(INPUT_FILE, data_only=True)
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

        # Print first 20 rows to understand structure
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
            if any(cell is not None for cell in row):
                print(f"Row {i + 1}: {row[:5]}...")  # Print first 5 cells

    wb.close()
    return wb.sheetnames


def extract_mapping_data():
    """Extract the indicator-to-code mapping from the Excel file."""
    print(f"\n{'=' * 60}")
    print("Extracting mapping data...")
    print(f"{'=' * 60}")

    # Read both sheets
    try:
        df_codes = pd.read_excel(INPUT_FILE, sheet_name='"Идеальные коды"', header=None)
        df_indicators = pd.read_excel(
            INPUT_FILE, sheet_name="Показатели и разрезы", header=None
        )

        print(f"\n--- Sheet 1 (Codes): Shape {df_codes.shape} ---")
        print(df_codes.head(15))

        print(f"\n--- Sheet 2 (Indicators): Shape {df_indicators.shape} ---")
        print(df_indicators.head(20))

        # Print column headers for sheet 2
        print(f"\nSheet 2 columns (first row):")
        print(df_indicators.iloc[0].tolist())

        return df_codes, df_indicators
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return None, None


def process_codes_sheet(df_codes):
    """Process the codes sheet to extract all ideal codes."""
    print(f"\n{'=' * 60}")
    print("Processing codes sheet...")
    print(f"{'=' * 60}")

    # The first row is header: ['Код ЦБСД', 'Комментарии']
    df_codes.columns = df_codes.iloc[0]
    df_codes = df_codes.iloc[1:].reset_index(drop=True)

    # Extract codes (first column)
    codes_df = df_codes[["Код ЦБСД", "Комментарии"]].copy()
    codes_df = codes_df.dropna(subset=["Код ЦБСД"])
    codes_df["Код ЦБСД"] = codes_df["Код ЦБСД"].astype(str).str.strip()

    print(f"Extracted {len(codes_df)} codes")
    print(f"Sample codes:")
    print(codes_df.head(10))

    return codes_df


def process_indicators_sheet(df_indicators):
    """Process the indicators sheet to extract indicator details."""
    print(f"\n{'=' * 60}")
    print("Processing indicators sheet...")
    print(f"{'=' * 60}")

    # The first row is header
    df_indicators.columns = df_indicators.iloc[0]
    df_indicators = df_indicators.iloc[1:].reset_index(drop=True)

    # Look for code column
    code_col = None
    name_col = None
    for col in df_indicators.columns:
        col_str = str(col).lower()
        if "код" in col_str:
            code_col = col
        if "наимен" in col_str or "детализ" in col_str:
            if name_col is None:  # Prefer first name column
                name_col = col

    print(f"Code column: {code_col}")
    print(f"Name column: {name_col}")

    # Extract relevant columns
    cols_to_keep = []
    if code_col:
        cols_to_keep.append(code_col)
    if name_col:
        cols_to_keep.append(name_col)

    # Also add frequency and depth if available
    for col in df_indicators.columns:
        col_str = str(col).lower()
        if "периодичн" in col_str or "глубин" in col_str:
            cols_to_keep.append(col)

    df_clean = df_indicators[cols_to_keep].copy()
    df_clean = df_clean.dropna(subset=[code_col] if code_col else [])

    print(f"Extracted {len(df_clean)} indicators")
    print(f"Sample indicators:")
    print(df_clean.head(10))

    return df_clean


def extract_registry(df_codes, df_indicators):
    """Extract the mapping registry by combining codes and indicators."""
    print(f"\n{'=' * 60}")
    print("Extracting registry...")
    print(f"{'=' * 60}")

    # Entry type 1: Codes with comments
    code_entries = []
    for _, row in df_codes.iterrows():
        code = str(row["Код ЦБСД"]).strip()
        comment = row.get("Комментарии", "")
        code_entries.append(
            {
                "indicator_code": code,
                "indicator_name": "",
                "comment": str(comment) if pd.notna(comment) else "",
                "source": "Идеальные коды",
            }
        )

    # Entry type 2: Indicators with their codes
    indicator_entries = []
    # Find the code column name
    code_col = None
    name_col = None
    for col in df_indicators.columns:
        col_str = str(col).lower()
        if "код" in col_str and code_col is None:
            code_col = col
        if ("наимен" in col_str or "детализ" in col_str) and name_col is None:
            name_col = col

    if code_col and name_col:
        for _, row in df_indicators.iterrows():
            code = row[code_col]
            name = row[name_col]
            if pd.notna(code):
                indicator_entries.append(
                    {
                        "indicator_code": str(code).strip(),
                        "indicator_name": str(name).strip(),
                        "comment": "",
                        "source": "Показатели и разрезы",
                    }
                )

    # Combine both
    registry_df = pd.DataFrame(code_entries + indicator_entries)

    # Merge entries that have same code - combine names from both sheets
    def combine_names(x):
        result = []
        for item in x:
            if pd.notna(item) and str(item) not in ["nan", ""]:
                result.append(str(item))
        return " | ".join(result) if result else ""

    def combine_sources(x):
        return ", ".join(sorted(set(str(s) for s in x if pd.notna(s))))

    def combine_comments(x):
        result = []
        for item in x:
            if pd.notna(item) and str(item) not in ["nan", ""]:
                result.append(str(item))
        return " | ".join(result) if result else ""

    registry_df = (
        registry_df.groupby("indicator_code")
        .agg(
            {
                "indicator_name": combine_names,
                "comment": combine_comments,
                "source": combine_sources,
            }
        )
        .reset_index()
    )

    print(f"Extracted {len(registry_df)} unique mappings")
    print(f"\nSample mappings:")
    print(registry_df.head(15))

    return registry_df


def check_for_weights(df_indicators):
    """Check if the file contains regional weights or aggregation logic."""
    print(f"\n{'=' * 60}")
    print("Checking for regional weights...")
    print(f"{'=' * 60}")

    weight_keywords = [
        "вес",
        "weight",
        "коэфф",
        "coeff",
        "агрег",
        "агрег",
        "регион",
        "region",
    ]

    found_weights = False
    weight_info = {}

    for col in df_indicators.columns:
        col_str = str(col).lower()
        if any(keyword in col_str for keyword in weight_keywords):
            print(f"Found weight-related column: {col}")
            found_weights = True
            weight_info[col] = df_indicators[col].dropna().head(10).tolist()

    return found_weights, weight_info


def save_results(registry, found_weights, weight_info, sheet_names):
    """Save the results to files."""
    print(f"\n{'=' * 60}")
    print("Saving results...")
    print(f"{'=' * 60}")

    # Ensure output directory exists
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    # Save registry
    registry.to_csv(OUTPUT_FILE, index=False, encoding="utf-8-sig")
    print(f"Registry saved to: {OUTPUT_FILE}")

    # Save metadata
    metadata = {
        "source_file": str(INPUT_FILE),
        "total_mappings": len(registry),
        "sheet_names": sheet_names,
        "has_weights": found_weights,
        "weight_info": {k: [str(v) for v in vals] for k, vals in weight_info.items()},
        "extraction_date": pd.Timestamp.now().isoformat(),
    }

    with open(METADATA_FILE, "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    print(f"Metadata saved to: {METADATA_FILE}")

    # Print summary
    print(f"\n{'=' * 60}")
    print("SUMMARY")
    print(f"{'=' * 60}")
    print(f"Total indicator mappings: {len(registry)}")
    print(f"Contains weights/aggregation logic: {found_weights}")
    print(f"Registry covers 100+ indicators: {'YES' if len(registry) >= 100 else 'NO'}")

    return len(registry) >= 100


def main():
    """Main execution function."""
    # Step 1: Analyze structure
    sheet_names = analyze_excel_structure()

    # Step 2: Extract data from both sheets
    df_codes, df_indicators = extract_mapping_data()
    if df_codes is None or df_indicators is None:
        print("Failed to extract data from Excel file")
        return False

    # Step 3: Process codes sheet
    codes_df = process_codes_sheet(df_codes)

    # Step 4: Process indicators sheet
    indicators_df = process_indicators_sheet(df_indicators)

    # Step 5: Extract registry
    registry = extract_registry(codes_df, indicators_df)

    # Step 6: Check for weights
    found_weights, weight_info = check_for_weights(indicators_df)

    # Step 7: Save results
    success = save_results(registry, found_weights, weight_info, sheet_names)

    print(f"\n{'=' * 60}")
    if success:
        print("✓ Task 125: COMPLETED successfully")
    else:
        print("✗ Task 125: Warning - Registry has < 100 indicators")
    print(f"{'=' * 60}")

    return success


if __name__ == "__main__":
    main()
