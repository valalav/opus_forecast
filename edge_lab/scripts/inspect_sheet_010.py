import pandas as pd
import openpyxl

# Load the workbook
wb_path = "/home/valalav/_projects/sirena-kbr/edge_lab/assets/charts/ОПР_статистика/Основная статистика ЮГУ.xlsx"
wb = openpyxl.load_workbook(wb_path)

# List all sheets
print("Available sheets:", wb.sheetnames)

# Load sheet '010'
if "010" in wb.sheetnames:
    ws = wb["010"]

    print("\n=== Sheet 010 Structure ===")

    # Print first 20 rows, first 10 columns
    print("\nFirst 20 rows, first 10 columns:")
    for row_idx in range(1, 21):
        row_data = []
        for col_idx in range(1, 11):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data.append(str(cell_value)[:30] if cell_value else "")
        print(f"Row {row_idx}: {row_data}")

    # Check cell (6, 1) for region 'Кбр'
    cell_6_1 = ws.cell(row=6, column=1).value
    print(f"\nCell (6, 1): {cell_6_1}")

    # Check row 10 for dates
    print("\nRow 10 (Dates):")
    row_10 = []
    for col_idx in range(1, 20):
        cell_value = ws.cell(row=10, column=col_idx).value
        row_10.append(str(cell_value)[:15] if cell_value else "")
    print(row_10)

    # Check column 3 for metric types
    print("\nColumn 3 (Metric Type) from row 11 onward:")
    for row_idx in range(11, 30):
        cell_value = ws.cell(row=row_idx, column=3).value
        print(f"Row {row_idx}, Col 3: {cell_value}")

else:
    print("Sheet '010' not found!")
