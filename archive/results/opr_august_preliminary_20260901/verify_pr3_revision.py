#!/usr/bin/env python3
"""Independently verify the preliminary-August PR3 workbook revision."""

from __future__ import annotations

import csv
import json
import zipfile
from math import prod
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "archive" / "results" / "opr_august_preliminary_20260901"
WORKBOOK = ROOT / "assets" / "06_2026_02_Прогноз.xlsx"
BACKUP = OUT_DIR / "06_2026_02_Прогноз.before_august_preliminary_revision.xlsx"
TRAJECTORY = OUT_DIR / "pr3_trajectory_revision.csv"
POLICY = ROOT / "data" / "send_ready_policy_trajectory.json"
FORECASTS = ROOT / "data" / "precomputed_forecasts.json"
MONTHLY = ROOT / "data" / "inflation_data.csv"


def yoy_index(values: dict[int, float], row: int) -> float:
    return round(prod(values[idx] for idx in range(row - 11, row + 1)) / 100**11, 2)


def main() -> None:
    failures: list[str] = []

    revised_formula = load_workbook(WORKBOOK, data_only=False, read_only=True)
    revised_values = load_workbook(WORKBOOK, data_only=True, read_only=True)
    backup = load_workbook(BACKUP, data_only=False, read_only=True)
    ws_formula = revised_formula["Прогноз"]
    ws_values = revised_values["Прогноз"]
    ws_backup = backup["Прогноз"]

    if revised_formula.sheetnames != backup.sheetnames or len(revised_formula.sheetnames) != 8:
        failures.append("workbook sheet structure changed")
    if float(ws_backup["E58"].value) != 100.50 or float(ws_backup["E59"].value) != 100.65:
        failures.append("backup does not contain the sent PR3 baseline")
    if float(ws_formula["E58"].value) != 100.00 or float(ws_formula["E59"].value) != 100.50:
        failures.append("revised August/September points are incorrect")

    unchanged_rows = range(60, 70)
    for row in unchanged_rows:
        if float(ws_formula.cell(row, 5).value) != float(ws_backup.cell(row, 5).value):
            failures.append(f"unexpected MoM change at row {row}")

    mom_values = {
        row: float(ws_formula.cell(row, 5).value)
        for row in range(3, 75)
        if ws_formula.cell(row, 5).value is not None
    }
    for row in range(58, 70):
        expected_formula = f"=ROUND(PRODUCT(E{row - 11}:E{row})/100^11,2)"
        if ws_formula.cell(row, 6).value != expected_formula:
            failures.append(f"YoY formula changed at F{row}")
        expected_yoy = yoy_index(mom_values, row)
        cached_yoy = float(ws_values.cell(row, 6).value)
        if cached_yoy != expected_yoy:
            failures.append(
                f"cached YoY mismatch at F{row}: expected {expected_yoy}, got {cached_yoy}"
            )

    with zipfile.ZipFile(WORKBOOK) as workbook_zip:
        if workbook_zip.testzip() is not None:
            failures.append("workbook ZIP container is corrupt")
        names = workbook_zip.namelist()
        if len([name for name in names if name.startswith("xl/charts/chart")]) != 3:
            failures.append("workbook no longer contains three charts")
        if "xl/externalLinks/externalLink1.xml" not in names:
            failures.append("workbook external link is missing")

    policy = json.loads(POLICY.read_text(encoding="utf-8"))
    forecasts = json.loads(FORECASTS.read_text(encoding="utf-8"))
    if policy["forecast_dates"] != forecasts["forecast_dates"]:
        failures.append("policy dates do not match the production forecast horizon")
    workbook_path = [float(ws_formula.cell(row, 5).value) - 100 for row in range(58, 70)]
    if len(policy["mom_pp"]) != len(workbook_path) or any(
        abs(policy_value - workbook_value) > 1e-9
        for policy_value, workbook_value in zip(policy["mom_pp"], workbook_path)
    ):
        failures.append("policy path does not match PR3 workbook values")

    with TRAJECTORY.open(encoding="utf-8", newline="") as source:
        rows = list(csv.DictReader(source))
    if len(rows) != 12 or rows[0]["new_mom_index"] != "100.00" or rows[1]["new_mom_index"] != "100.50":
        failures.append("trajectory revision CSV is inconsistent")

    monthly = pd.read_csv(MONTHLY, sep=";", decimal=",", encoding="utf-8-sig")
    latest_official = pd.to_datetime(monthly["Date"], dayfirst=True).max()
    if latest_official.strftime("%Y-%m") != "2026-07":
        failures.append("official monthly data unexpectedly include August")

    if failures:
        print("PR3 REVISION VERIFICATION: FAIL")
        for failure in failures:
            print(f"- {failure}")
        raise SystemExit(1)

    print("PR3 REVISION VERIFICATION: PASS")
    print("- sent baseline backup: August 100.50, September 100.65")
    print("- revised workbook: August 100.00, September 100.50")
    print("- October 2026 through July 2027 MoM path unchanged")
    print("- formulas and cached YoY values verified for F58:F69")
    print("- 8 sheets, 3 charts, and external link preserved")
    print("- dashboard policy path aligned to precomputed forecast dates")
    print("- official monthly source still ends at July 2026")


if __name__ == "__main__":
    main()
