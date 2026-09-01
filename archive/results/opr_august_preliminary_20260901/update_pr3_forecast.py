#!/usr/bin/env python3
"""Apply the 2026-09-01 preliminary-August revision to the live PR3 workbook."""

from __future__ import annotations

import csv
import shutil
from datetime import datetime
from math import prod
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = ROOT / "assets" / "06_2026_02_Прогноз.xlsx"
OUT_DIR = ROOT / "archive" / "results" / "opr_august_preliminary_20260901"
BACKUP = OUT_DIR / "06_2026_02_Прогноз.before_august_preliminary_revision.xlsx"
TRAJECTORY = OUT_DIR / "pr3_trajectory_revision.csv"

EXPECTED = {
    58: (datetime(2026, 8, 1), 100.50),
    59: (datetime(2026, 9, 1), 100.65),
}
REVISED = {
    58: 100.00,
    59: 100.50,
}


def yoy_index(values: dict[int, float], row: int) -> float:
    """Return the workbook's rolling 12-month YoY index for one row."""
    return round(prod(values[idx] for idx in range(row - 11, row + 1)) / 100**11, 2)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(WORKBOOK, data_only=False, keep_links=True)
    sheet = workbook["Прогноз"]

    for row, (expected_date, expected_value) in EXPECTED.items():
        actual_date = sheet.cell(row, 1).value
        actual_value = float(sheet.cell(row, 5).value)
        if actual_date != expected_date or actual_value != expected_value:
            raise RuntimeError(
                f"Unexpected PR3 baseline at row {row}: "
                f"date={actual_date!r}, value={actual_value!r}"
            )

    if BACKUP.exists():
        raise FileExistsError(f"Backup already exists: {BACKUP}")
    shutil.copy2(WORKBOOK, BACKUP)

    old_values = {
        row: float(sheet.cell(row, 5).value)
        for row in range(3, 75)
        if sheet.cell(row, 5).value is not None
    }
    new_values = old_values.copy()
    new_values.update(REVISED)

    revision_date = datetime(2026, 9, 1)
    sheet["E1"] = revision_date
    sheet["F1"] = revision_date
    for row, value in REVISED.items():
        sheet.cell(row, 5).value = value

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.modified = revision_date
    workbook.save(WORKBOOK)

    with TRAJECTORY.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(
            [
                "month",
                "old_mom_index",
                "new_mom_index",
                "old_yoy_index",
                "new_yoy_index",
                "yoy_change_pp",
                "status",
            ]
        )
        for row in range(58, 70):
            old_yoy = yoy_index(old_values, row)
            new_yoy = yoy_index(new_values, row)
            month = sheet.cell(row, 1).value.strftime("%Y-%m")
            status = "preliminary" if row == 58 else "forecast"
            writer.writerow(
                [
                    month,
                    f"{old_values[row]:.2f}",
                    f"{new_values[row]:.2f}",
                    f"{old_yoy:.2f}",
                    f"{new_yoy:.2f}",
                    f"{new_yoy - old_yoy:.2f}",
                    status,
                ]
            )

    print(f"Updated: {WORKBOOK}")
    print(f"Backup: {BACKUP}")
    print(f"Trajectory: {TRAJECTORY}")


if __name__ == "__main__":
    main()
