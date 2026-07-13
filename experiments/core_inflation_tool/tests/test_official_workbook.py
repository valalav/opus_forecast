from datetime import datetime

import openpyxl
import pytest

from experiments.core_inflation_tool.core_inflation.official_workbook import (
    OFFICIAL_COMPONENT_ROWS,
    extract_official_results,
    validate_official_formulas,
)


def test_extract_official_results_reproduces_excel_median_exactly():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Результаты"
    sheet.cell(1, 3, "01.05.2026")
    values = [float(value) for value in range(1, len(OFFICIAL_COMPONENT_ROWS) + 1)]
    for row, value in zip(OFFICIAL_COMPONENT_ROWS, values, strict=True):
        sheet.cell(row, 3, value)
    expected = 8.5
    sheet.cell(4, 3, expected)
    sheet.cell(5, 3, 6.0)

    series, components, differences = extract_official_results(workbook)

    assert series.loc[0, "date"] == datetime(2026, 5, 1).date().isoformat()
    assert series.loc[0, "official_stable_inflation_saar"] == pytest.approx(expected)
    assert series.loc[0, "reproduced_median_saar"] == pytest.approx(expected)
    assert differences == pytest.approx([0.0])
    assert len(components) == len(OFFICIAL_COMPONENT_ROWS)


def test_extract_official_results_adds_transparent_analytics():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Результаты"
    values = [float(value) for value in range(1, len(OFFICIAL_COMPONENT_ROWS) + 1)]
    for column, (period, offset, reverse) in enumerate(
        [("01.03.2026", 0.0, False), ("01.04.2026", 1.0, True), ("01.05.2026", 2.0, True)],
        start=3,
    ):
        sheet.cell(1, column, period)
        month_values = list(reversed(values)) if reverse else values
        for row, value in zip(OFFICIAL_COMPONENT_ROWS, month_values, strict=True):
            sheet.cell(row, 2, f"Показатель {row}")
            sheet.cell(row, column, value + offset)
        sheet.cell(4, column, 8.5 + offset)

    series, components, _ = extract_official_results(workbook)

    expected_monthly = ((1.0 + series["official_stable_inflation_saar"] / 100.0) ** (1.0 / 12.0) - 1.0) * 100.0
    expected_3mma_saar = ((1.0 + expected_monthly.mean() / 100.0) ** 12 - 1.0) * 100.0
    assert series["official_stable_inflation_mom_sa"].tolist() == pytest.approx(expected_monthly.tolist())
    assert series.loc[2, "official_stable_inflation_3mma_saar"] == pytest.approx(expected_3mma_saar)
    assert series["median_pair_changed"].tolist() == [False, True, False]
    assert series.loc[0, "median_lower_saar"] == pytest.approx(8.0)
    assert series.loc[0, "median_upper_saar"] == pytest.approx(9.0)
    assert series.loc[0, "component_min_saar"] == pytest.approx(1.0)
    assert series.loc[0, "component_max_saar"] == pytest.approx(16.0)
    assert components["component_label"].notna().all()


def test_validate_official_formulas_checks_references_and_random_tiebreaker():
    workbook = openpyxl.Workbook()
    results = workbook.active
    results.title = "Результаты"
    results.cell(1, 3, "01.05.2026")
    results.cell(4, 3, "=IFERROR(MEDIAN(C7:C10,C13:C14,C16:C17,C18,C22:C27,C2),\"\")")
    month = workbook.create_sheet("01.05.2026")
    month.cell(137, 2, "=STDEV.S(B2:B4)+RAND()/10000")

    checks, matches, random_count = validate_official_formulas(workbook)

    assert (checks, matches, random_count) == (1, 1, 1)
