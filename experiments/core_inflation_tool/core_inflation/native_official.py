"""Native Linux implementation of the approved stable-inflation methodology.

The calculator reads source data only. It never asks Excel or LibreOffice to
recalculate formulas. An approved workbook may be supplied separately as a
read-only validation reference.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
from pathlib import Path
from statistics import median
from typing import Any, Iterable

import openpyxl
import pandas as pd

from .loaders import load_long_item_indices, load_weights
from .official_workbook import OFFICIAL_COMPONENT_ROWS, extract_official_results

COMPONENT_CODES = frozenset(set(range(11, 45)) | set(range(46, 56)) | {67})
FIXED_ADMIN_CODES = frozenset({14, 33, 42})
SEVEN_VOLATILE_CODES = frozenset({18, 33, 36, 39, 42, 48, 52})
CREDIT_ITEM_CODE = 479
CREDIT_ITEM_NAME = "Плата за пользование потребительским кредитом (процентная ставка в стоимостном выражении), руб."
OTHER_SERVICES_CODE = 55
BASE_CPI_CODE = 5
FOOD_EX_FRUIT_REGSA_CODE = 7
FOOD_EX_FRUIT_WEIGHT_CODE = 8
CALCULATION_START = pd.Timestamp("2021-01-01")
TOLERANCE = 1e-9


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_record(path: Path) -> dict[str, object]:
    return {
        "path": str(path.resolve()),
        "size_bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def annualize_index(index_value: float) -> float:
    """Convert a monthly growth index around 100 into SAAR percent."""
    if not math.isfinite(index_value) or index_value <= 0:
        raise ValueError(f"monthly index must be finite and positive, got {index_value!r}")
    return index_value**12 / 100.0**11 - 100.0


def deannualize_saar(annual_rate: pd.Series) -> pd.Series:
    """Return the monthly percent rate implied by a SAAR series."""
    return ((1.0 + annual_rate / 100.0) ** (1.0 / 12.0) - 1.0) * 100.0


def _trim_weight(values: pd.Series, weights: pd.Series, amount: float, *, reverse: bool) -> pd.Series:
    remaining = weights.astype(float).copy()
    left = float(amount)
    positions: Iterable[int] = reversed(range(len(remaining))) if reverse else range(len(remaining))
    for position in positions:
        removed = min(float(remaining.iloc[position]), left)
        remaining.iloc[position] -= removed
        left -= removed
        if left <= 1e-15:
            break
    if left > 1e-12:
        raise ValueError(f"cannot trim {amount:.6f}; only {amount - left:.6f} weight is available")
    return remaining


def weighted_trimmed_index(values: pd.Series, weights: pd.Series, excluded_share: float) -> float:
    """Return an index after removing equal weighted tails with partial boundaries."""
    if not 0.0 <= excluded_share < 1.0:
        raise ValueError("excluded_share must be in [0, 1)")
    ordered = pd.DataFrame({"value": values, "weight": weights}).dropna().sort_values(
        "value", kind="stable"
    )
    remaining = _trim_weight(ordered["value"], ordered["weight"], excluded_share / 2.0, reverse=False)
    remaining = _trim_weight(ordered["value"], remaining, excluded_share / 2.0, reverse=True)
    return float((ordered["value"] * remaining).sum() / remaining.sum())


def volatility_exclusion_index(
    current: pd.Series,
    volatility: pd.Series,
    weights: pd.Series,
    excluded_share: float,
    *,
    forced_codes: frozenset[int] = frozenset(),
) -> float:
    """Exclude the highest-volatility weighted share with a partial boundary."""
    if not 0.0 <= excluded_share < 1.0:
        raise ValueError("excluded_share must be in [0, 1)")
    rows = pd.DataFrame(
        {
            "code": current.index.astype(int),
            "value": current.values,
            "volatility": volatility.reindex(current.index).values,
            "weight": weights.reindex(current.index).values,
        }
    ).dropna()
    rows["forced"] = rows["code"].isin(forced_codes)
    rows = rows.sort_values(
        ["forced", "volatility", "code"], ascending=[False, False, True], kind="stable"
    ).reset_index(drop=True)
    remaining = _trim_weight(rows["value"], rows["weight"], excluded_share, reverse=False)
    return float((rows["value"] * remaining).sum() / remaining.sum())


def _read_regsa(regsa_path: Path) -> tuple[pd.DataFrame, dict[int, str]]:
    workbook = openpyxl.load_workbook(regsa_path, read_only=True, data_only=True)
    if "mom_sa" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("RegSA workbook must contain a 'mom_sa' sheet")
    sheet = workbook["mom_sa"]
    sheet_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if len(sheet_rows) < 5:
        raise ValueError("RegSA 'mom_sa' sheet has no data rows")
    dates = [
        (position, pd.Timestamp(value))
        for position, value in enumerate(sheet_rows[3][3:], start=3)
        if value is not None
    ]
    rows: list[dict[str, object]] = []
    names: dict[int, str] = {}
    for source_row in sheet_rows[4:]:
        code = source_row[1] if len(source_row) > 1 else None
        name = source_row[2] if len(source_row) > 2 else None
        if code is None or name is None:
            continue
        item_code = int(code)
        names[item_code] = str(name)
        for position, date in dates:
            value = source_row[position] if position < len(source_row) else None
            if isinstance(value, (int, float)) and math.isfinite(float(value)):
                rows.append({"date": date, "item_code": item_code, "index_value": float(value)})
    frame = pd.DataFrame(rows).sort_values(["date", "item_code"]).reset_index(drop=True)
    return frame, names


def _read_credit_from_workbook(path: Path) -> pd.Series:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=True)
    if "ИПЦ_все" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("raw workbook must contain an 'ИПЦ_все' sheet")
    sheet = workbook["ИПЦ_все"]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        credit_column = header.index(CREDIT_ITEM_NAME)
    except ValueError as error:
        workbook.close()
        raise ValueError(f"'ИПЦ_все' does not contain {CREDIT_ITEM_NAME!r}") from error
    values: dict[pd.Timestamp, float] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None or credit_column >= len(row):
            continue
        value = row[credit_column]
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            values[pd.Timestamp(row[0])] = 100.0 + float(value)
    workbook.close()
    return pd.Series(values, dtype=float, name="credit_index").sort_index()


def _read_credit_from_csv(path: Path, region_code: int) -> pd.Series:
    frame = load_long_item_indices(path, region_code=region_code)
    credit = frame.loc[frame["item_code"].astype(int).eq(CREDIT_ITEM_CODE), ["date", "mom_index"]]
    if credit.empty:
        raise ValueError(f"{path} contains no item {CREDIT_ITEM_CODE} for region {region_code}")
    return credit.drop_duplicates("date").set_index("date")["mom_index"].sort_index()


def read_credit_index(path: Path, region_code: int) -> pd.Series:
    """Read the credit-price monthly index from CSV or an input workbook."""
    if path.suffix.lower() == ".csv":
        return _read_credit_from_csv(path, region_code)
    return _read_credit_from_workbook(path)


def prepare_inputs(
    regsa_path: Path,
    weights_path: Path,
    raw_mom_path: Path,
    *,
    region_code: int = 7,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[int, str], dict[str, object]]:
    """Load and validate the component panel and annual vertical weights."""
    regsa, names = _read_regsa(regsa_path)
    weights = load_weights(weights_path, region_code=region_code, weight_column="Weight_vertical").rename(
        columns={"weight_year": "year"}
    )
    credit = read_credit_index(raw_mom_path, region_code)

    component_panel = regsa.loc[regsa["item_code"].isin(COMPONENT_CODES)].copy()
    available_codes = set(component_panel["item_code"].astype(int).unique())
    if available_codes != COMPONENT_CODES:
        missing = sorted(COMPONENT_CODES - available_codes)
        extra = sorted(available_codes - COMPONENT_CODES)
        raise ValueError(f"RegSA component hierarchy mismatch; missing={missing}, extra={extra}")

    relevant_dates = component_panel["date"].drop_duplicates().sort_values()
    missing_credit = [date for date in relevant_dates if date >= CALCULATION_START and date not in credit.index]
    if missing_credit:
        raise ValueError(
            f"credit input is missing {len(missing_credit)} calculation months; "
            f"first={missing_credit[0].date()}, last={missing_credit[-1].date()}"
        )

    required_weight_codes = COMPONENT_CODES | {
        CREDIT_ITEM_CODE,
        BASE_CPI_CODE,
        FOOD_EX_FRUIT_WEIGHT_CODE,
    }
    weight_years = set(component_panel.loc[component_panel["date"] >= CALCULATION_START, "date"].dt.year)
    for year in sorted(weight_years):
        year_weights = weights.loc[weights["year"].eq(year)].set_index("item_code")["weight"]
        missing = sorted(required_weight_codes - set(year_weights.index.astype(int)))
        if missing:
            raise ValueError(f"vertical weights for {year} are missing codes {missing}")
        component_sum = float(year_weights.reindex(sorted(COMPONENT_CODES)).sum())
        if abs(component_sum - 1.0) > 1e-9:
            raise ValueError(f"component weights for {year} sum to {component_sum:.12f}, expected 1")

    for row_index, row in component_panel.loc[component_panel["item_code"].eq(OTHER_SERVICES_CODE)].iterrows():
        date = pd.Timestamp(row["date"])
        if date not in credit.index:
            continue
        year_weights = weights.loc[weights["year"].eq(date.year)].set_index("item_code")["weight"]
        ratio = float(year_weights.loc[CREDIT_ITEM_CODE] / year_weights.loc[OTHER_SERVICES_CODE])
        component_panel.loc[row_index, "index_value"] = (
            float(row["index_value"]) - ratio * float(credit.loc[date])
        ) / (1.0 - ratio)

    diagnostics: dict[str, object] = {
        "region_code": region_code,
        "component_count": len(COMPONENT_CODES),
        "component_dates": len(relevant_dates),
        "component_start": relevant_dates.iloc[0].date().isoformat(),
        "component_end": relevant_dates.iloc[-1].date().isoformat(),
        "credit_start": credit.index.min().date().isoformat(),
        "credit_end": credit.index.max().date().isoformat(),
        "weight_years": sorted(weight_years),
        "maximum_component_weight_sum_error": max(
            abs(
                float(
                    weights.loc[
                        weights["year"].eq(year) & weights["item_code"].isin(COMPONENT_CODES), "weight"
                    ].sum()
                )
                - 1.0
            )
            for year in weight_years
        ),
    }
    return component_panel, weights, regsa, names, diagnostics


def _weighted_exclusion_index(
    current: pd.Series,
    weights: pd.Series,
    excluded_codes: frozenset[int],
) -> float:
    keep = current.index.difference(list(excluded_codes))
    retained_weights = weights.reindex(keep)
    return float((current.reindex(keep) * retained_weights).sum() / retained_weights.sum())


def calculate_native_series(
    component_panel: pd.DataFrame,
    weights: pd.DataFrame,
    regsa: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Calculate all 16 approved estimates and their monthly median."""
    panel = component_panel.pivot(index="date", columns="item_code", values="index_value").sort_index()
    aggregates = regsa.pivot(index="date", columns="item_code", values="index_value").sort_index()
    dates = panel.index[panel.index >= CALCULATION_START]
    series_rows: list[dict[str, object]] = []
    component_rows: list[dict[str, object]] = []

    for date in dates:
        history = panel.loc[:date]
        if len(history) < 24:
            continue
        current = history.loc[date].dropna()
        if set(current.index.astype(int)) != COMPONENT_CODES:
            raise ValueError(f"{date.date()}: expected {len(COMPONENT_CODES)} component values, got {len(current)}")
        year_weights = weights.loc[
            weights["year"].eq(date.year) & weights["item_code"].isin(COMPONENT_CODES)
        ].set_index("item_code")["weight"]
        estimates: dict[str, float] = {}

        for share in (0.20, 0.40, 0.60, 0.80):
            estimates[f"trimmed_{int(share * 100)}"] = annualize_index(
                weighted_trimmed_index(current, year_weights, share)
            )

        for window in (3, 24):
            volatility = history.tail(window).std(ddof=1)
            for share in (0.10, 0.20):
                estimates[f"volatile_{window}m_ex_{int(share * 100)}"] = annualize_index(
                    volatility_exclusion_index(current, volatility, year_weights, share)
                )

        volatility_24m = history.tail(24).std(ddof=1)
        fixed_weight = float(year_weights.reindex(sorted(FIXED_ADMIN_CODES)).sum())
        estimates["volatile_24m_ex_30"] = annualize_index(
            volatility_exclusion_index(
                current,
                volatility_24m,
                year_weights,
                0.30,
                forced_codes=FIXED_ADMIN_CODES,
            )
        )
        estimates["volatile_24m_ex_admin_and_20"] = annualize_index(
            volatility_exclusion_index(
                current,
                volatility_24m,
                year_weights,
                fixed_weight + 0.20,
                forced_codes=FIXED_ADMIN_CODES,
            )
        )

        aggregate = aggregates.loc[date]
        estimates["base_inflation"] = annualize_index(float(aggregate.loc[BASE_CPI_CODE]))
        all_year_weights = weights.loc[weights["year"].eq(date.year)].set_index("item_code")["weight"]
        food_share_in_base = float(
            all_year_weights.loc[FOOD_EX_FRUIT_WEIGHT_CODE] / all_year_weights.loc[BASE_CPI_CODE]
        )
        base_ex_food_index = (
            float(aggregate.loc[BASE_CPI_CODE])
            - food_share_in_base * float(aggregate.loc[FOOD_EX_FRUIT_REGSA_CODE])
        ) / (1.0 - food_share_in_base)
        estimates["base_inflation_ex_food"] = annualize_index(base_ex_food_index)

        exclusions = {
            "cpi_ex_fruit_fuel_housing": FIXED_ADMIN_CODES,
            "cpi_ex_housing": frozenset({14}),
            "cpi_ex_fruit_vegetables": frozenset({33}),
            "cpi_ex_seven_volatile": SEVEN_VOLATILE_CODES,
        }
        for key, excluded_codes in exclusions.items():
            estimates[key] = annualize_index(
                _weighted_exclusion_index(current, year_weights, excluded_codes)
            )

        if set(estimates) != set(OFFICIAL_COMPONENT_ROWS.values()):
            missing = sorted(set(OFFICIAL_COMPONENT_ROWS.values()) - set(estimates))
            extra = sorted(set(estimates) - set(OFFICIAL_COMPONENT_ROWS.values()))
            raise AssertionError(f"official estimate set mismatch; missing={missing}, extra={extra}")
        official = float(median(estimates.values()))
        series_rows.append(
            {
                "date": date.date().isoformat(),
                "native_stable_inflation_saar": official,
                "estimate_count": len(estimates),
            }
        )
        for component, value in estimates.items():
            component_rows.append(
                {
                    "date": date.date().isoformat(),
                    "component": component,
                    "value_saar": value,
                }
            )

    series = pd.DataFrame(series_rows)
    if not series.empty:
        series["native_stable_inflation_mom_sa"] = deannualize_saar(series["native_stable_inflation_saar"])
        series["native_stable_inflation_3mma_mom_sa"] = series["native_stable_inflation_mom_sa"].rolling(
            3, min_periods=3
        ).mean()
        series["native_stable_inflation_3mma_saar"] = (
            (1.0 + series["native_stable_inflation_3mma_mom_sa"] / 100.0) ** 12 - 1.0
        ) * 100.0
    return series, pd.DataFrame(component_rows)


def validate_against_reference(
    series: pd.DataFrame,
    components: pd.DataFrame,
    reference_path: Path,
) -> tuple[dict[str, object], pd.DataFrame]:
    """Compare native results with cached approved-workbook values without recalculation."""
    workbook = openpyxl.load_workbook(
        reference_path, read_only=True, data_only=True, keep_vba=True, keep_links=True
    )
    official_series, official_components, _ = extract_official_results(workbook)
    workbook.close()
    component_comparison = components.merge(
        official_components[["date", "component", "value_saar"]],
        on=["date", "component"],
        how="outer",
        suffixes=("_native", "_excel"),
    )
    component_comparison["difference"] = (
        component_comparison["value_saar_native"] - component_comparison["value_saar_excel"]
    )
    series_comparison = series.merge(
        official_series[["date", "official_stable_inflation_saar"]], on="date", how="outer"
    )
    series_comparison["difference"] = (
        series_comparison["native_stable_inflation_saar"]
        - series_comparison["official_stable_inflation_saar"]
    )
    component_complete = not component_comparison[["value_saar_native", "value_saar_excel"]].isna().any().any()
    series_complete = not series_comparison[
        ["native_stable_inflation_saar", "official_stable_inflation_saar"]
    ].isna().any().any()
    component_max = float(component_comparison["difference"].abs().max())
    series_max = float(series_comparison["difference"].abs().max())
    validation = {
        "component_cells": len(component_comparison),
        "component_exact_matches": int((component_comparison["difference"].abs() <= TOLERANCE).sum()),
        "component_max_abs_difference": component_max,
        "official_months": len(series_comparison),
        "official_exact_matches": int((series_comparison["difference"].abs() <= TOLERANCE).sum()),
        "official_max_abs_difference": series_max,
        "component_coverage_complete": component_complete,
        "official_coverage_complete": series_complete,
        "tolerance": TOLERANCE,
        "passed": component_complete
        and series_complete
        and component_max <= TOLERANCE
        and series_max <= TOLERANCE,
    }
    return validation, component_comparison


def render_report(series: pd.DataFrame, validation: dict[str, object] | None) -> str:
    lines = [
        "# Нативный расчёт утверждённой устойчивой инфляции",
        "",
        "Все 16 оценок и итоговая медиана рассчитаны Python-кодом. Excel/LibreOffice не исполнялись.",
        "",
        "## Последние шесть месяцев",
        "",
        "| месяц | нативный SAAR | SA м/м | 3MMA SAAR |",
        "|---|---:|---:|---:|",
    ]
    for row in series.tail(6).itertuples(index=False):
        lines.append(
            f"| {row.date[:7]} | {row.native_stable_inflation_saar:.6f}% | "
            f"{row.native_stable_inflation_mom_sa:.6f}% | "
            f"{row.native_stable_inflation_3mma_saar:.6f}% |"
        )
    lines.extend(["", "## Валидация", ""])
    if validation is None:
        lines.append("Эталонная книга не передана; выполнены только входные контроли.")
    else:
        lines.extend(
            [
                f"- 16 показателей: {validation['component_exact_matches']}/{validation['component_cells']} совпали в допуске {validation['tolerance']:.0e}.",
                f"- Итоговая медиана: {validation['official_exact_matches']}/{validation['official_months']} месяцев совпали в допуске.",
                f"- Максимальная разница показателей: {validation['component_max_abs_difference']:.12g} п.п.",
                f"- Максимальная разница медианы: {validation['official_max_abs_difference']:.12g} п.п.",
                f"- Статус: {'PASS' if validation['passed'] else 'FAIL'}.",
            ]
        )
    lines.extend(
        [
            "",
            "## Граница расчёта",
            "",
            "`RegSA.xlsx` используется только как источник уже сезонно очищенных месячных индексов. "
            "Остаточный агрегат «Другие услуги» строится нативно с исключением платы за потребительский кредит. "
            "Ранжирование, weighted trimming, исключение волатильных компонент, перенормировка весов, "
            "аннуализация и итоговая медиана выполняются Python-кодом.",
            "",
        ]
    )
    return "\n".join(lines)


def run(
    regsa_path: Path,
    weights_path: Path,
    raw_mom_path: Path,
    output_path: Path,
    *,
    region_code: int = 7,
    reference_path: Path | None = None,
) -> dict[str, Path]:
    component_panel, weights, regsa, names, input_diagnostics = prepare_inputs(
        regsa_path, weights_path, raw_mom_path, region_code=region_code
    )
    series, components = calculate_native_series(component_panel, weights, regsa)
    validation: dict[str, object] | None = None
    comparison = pd.DataFrame()
    if reference_path is not None:
        validation, comparison = validate_against_reference(series, components, reference_path)

    output_path.mkdir(parents=True, exist_ok=True)
    paths = {
        "series": output_path / "native_official_stable_inflation_series.csv",
        "components": output_path / "native_official_stable_inflation_components.csv",
        "diagnostics": output_path / "native_official_diagnostics.json",
        "sources": output_path / "native_official_source_manifest.json",
        "report": output_path / "native_official_report.md",
    }
    if reference_path is not None:
        paths["comparison"] = output_path / "native_official_comparison.csv"
        paths["validation"] = output_path / "native_official_validation.json"

    source_paths = {"regsa": regsa_path, "weights": weights_path, "raw_mom": raw_mom_path}
    if reference_path is not None:
        source_paths["reference"] = reference_path
    manifest = {name: _source_record(path) for name, path in source_paths.items()}
    diagnostics: dict[str, Any] = input_diagnostics | {
        "calculated_months": len(series),
        "calculated_component_cells": len(components),
        "component_names": {str(code): names.get(code, str(code)) for code in sorted(COMPONENT_CODES)},
        "validation_passed": None if validation is None else validation["passed"],
    }

    series.to_csv(paths["series"], index=False)
    components.to_csv(paths["components"], index=False)
    paths["diagnostics"].write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    paths["sources"].write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    paths["report"].write_text(render_report(series, validation), encoding="utf-8")
    if reference_path is not None and validation is not None:
        comparison.to_csv(paths["comparison"], index=False)
        paths["validation"].write_text(
            json.dumps(validation, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
    return paths


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Calculate the approved stable-inflation methodology natively.")
    parser.add_argument("--regsa", required=True, type=Path, help="RegSA.xlsx source with the mom_sa sheet")
    parser.add_argument("--weights", required=True, type=Path, help="ACCDB-derived weights CSV")
    parser.add_argument(
        "--raw-mom",
        required=True,
        type=Path,
        help="Raw indices CSV or an input workbook containing ИПЦ_все",
    )
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--region-code", type=int, default=7)
    parser.add_argument(
        "--reference-workbook",
        type=Path,
        help="Optional approved workbook used only to validate cached results",
    )
    args = parser.parse_args(argv)
    paths = run(
        args.regsa,
        args.weights,
        args.raw_mom,
        args.output,
        region_code=args.region_code,
        reference_path=args.reference_workbook,
    )
    for name, path in paths.items():
        print(f"{name}: {path}")
    if "validation" in paths:
        validation = json.loads(paths["validation"].read_text(encoding="utf-8"))
        return 0 if validation["passed"] else 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
