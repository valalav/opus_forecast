"""Extract and validate the approved stable-inflation Excel methodology."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
import hashlib
import json
from pathlib import Path
from statistics import median
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd


RESULT_SHEET = "Результаты"
OFFICIAL_ROW = 4
OFFICIAL_COMPONENT_ROWS = {
    2: "volatile_24m_ex_30",
    7: "trimmed_20",
    8: "trimmed_40",
    9: "trimmed_60",
    10: "trimmed_80",
    13: "volatile_3m_ex_10",
    14: "volatile_3m_ex_20",
    16: "volatile_24m_ex_10",
    17: "volatile_24m_ex_20",
    18: "volatile_24m_ex_admin_and_20",
    22: "base_inflation",
    23: "base_inflation_ex_food",
    24: "cpi_ex_fruit_fuel_housing",
    25: "cpi_ex_housing",
    26: "cpi_ex_fruit_vegetables",
    27: "cpi_ex_seven_volatile",
}
MONTH_SHEET_ROWS = {
    2: 205,
    7: 126,
    8: 127,
    9: 128,
    10: 129,
    13: 150,
    14: 151,
    16: 171,
    17: 172,
    18: 206,
    20: 185,
}


@dataclass(frozen=True)
class ValidationSummary:
    official_months: int
    median_exact_matches: int
    median_max_abs_diff: float
    official_formula_checks: int
    official_formula_exact_matches: int
    random_tiebreaker_formula_count: int
    monthly_reference_checks: int
    monthly_reference_exact_matches: int
    regsa_checks: int
    regsa_exact_matches: int
    database_yoy_checks: int
    database_yoy_exact_matches: int
    mom_yoy_comparable_cells: int
    mom_yoy_identical_cells: int

    @property
    def passed(self) -> bool:
        return (
            self.official_months > 0
            and self.median_exact_matches == self.official_months
            and self.monthly_reference_checks == self.monthly_reference_exact_matches
            and self.regsa_checks == self.regsa_exact_matches
            and self.official_formula_checks == self.official_formula_exact_matches
            and self.database_yoy_checks > 0
            and self.database_yoy_checks == self.database_yoy_exact_matches
            and self.mom_yoy_comparable_cells > 0
            and self.mom_yoy_identical_cells < self.mom_yoy_comparable_cells
        )


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _is_blank(value: object) -> bool:
    return value is None or value == ""


def _parse_period(value: object) -> pd.Timestamp | None:
    if isinstance(value, datetime):
        return pd.Timestamp(value)
    if isinstance(value, str):
        parsed = pd.to_datetime(value, format="%d.%m.%Y", errors="coerce")
        return None if pd.isna(parsed) else pd.Timestamp(parsed)
    return None


def _sheet_values(workbook: openpyxl.Workbook, sheet_name: str) -> list[list[Any]]:
    return [[cell.value for cell in row] for row in workbook[sheet_name].iter_rows()]



def _annualize_monthly(monthly_rate: pd.Series) -> pd.Series:
    return ((1.0 + monthly_rate / 100.0) ** 12 - 1.0) * 100.0


def _deannualize_saar(annual_rate: pd.Series) -> pd.Series:
    return ((1.0 + annual_rate / 100.0) ** (1.0 / 12.0) - 1.0) * 100.0


def extract_official_results(workbook: openpyxl.Workbook) -> tuple[pd.DataFrame, pd.DataFrame, list[float]]:
    sheet = workbook[RESULT_SHEET]
    series_rows: list[dict[str, object]] = []
    component_rows: list[dict[str, object]] = []
    median_differences: list[float] = []

    for column in range(3, sheet.max_column + 1):
        period = _parse_period(sheet.cell(1, column).value)
        official = sheet.cell(OFFICIAL_ROW, column).value
        if period is None or not _is_number(official):
            continue

        components: list[tuple[str, str, float]] = []
        for row, key in OFFICIAL_COMPONENT_ROWS.items():
            value = sheet.cell(row, column).value
            if not _is_number(value):
                continue
            numeric = float(value)
            label = " ".join(str(sheet.cell(row, 2).value or key).split())
            components.append((key, label, numeric))
            component_rows.append(
                {
                    "date": period.date().isoformat(),
                    "component": key,
                    "component_label": label,
                    "result_row": row,
                    "value_saar": numeric,
                }
            )
        if len(components) != len(OFFICIAL_COMPONENT_ROWS):
            raise ValueError(f"{period.date()}: expected {len(OFFICIAL_COMPONENT_ROWS)} official components, got {len(components)}")

        ordered = sorted(components, key=lambda item: (item[2], item[0]))
        values = [value for _, _, value in ordered]
        reproduced = float(median(values))
        difference = float(official) - reproduced
        median_differences.append(difference)
        headline = sheet.cell(5, column).value
        series_rows.append(
            {
                "date": period.date().isoformat(),
                "official_stable_inflation_saar": float(official),
                "reproduced_median_saar": reproduced,
                "difference": difference,
                "headline_sa_saar": float(headline) if _is_number(headline) else None,
                "component_count": len(components),
                "component_min_saar": min(values),
                "component_q1_saar": float(pd.Series(values).quantile(0.25)),
                "component_q3_saar": float(pd.Series(values).quantile(0.75)),
                "component_max_saar": max(values),
                "component_iqr_saar": float(pd.Series(values).quantile(0.75) - pd.Series(values).quantile(0.25)),
                "median_lower_component": ordered[7][0],
                "median_lower_label": ordered[7][1],
                "median_lower_saar": ordered[7][2],
                "median_upper_component": ordered[8][0],
                "median_upper_label": ordered[8][1],
                "median_upper_saar": ordered[8][2],
            }
        )

    series = pd.DataFrame(series_rows).sort_values("date").reset_index(drop=True)
    if not series.empty:
        series["official_stable_inflation_mom_sa"] = _deannualize_saar(series["official_stable_inflation_saar"])
        series["official_stable_inflation_3mma_mom_sa"] = (
            series["official_stable_inflation_mom_sa"].rolling(3, min_periods=3).mean()
        )
        series["official_stable_inflation_3mma_saar"] = _annualize_monthly(
            series["official_stable_inflation_3mma_mom_sa"]
        )
        pair = series["median_lower_component"] + "|" + series["median_upper_component"]
        series["median_pair_changed"] = pair.ne(pair.shift()).fillna(False)
        series.loc[series.index[0], "median_pair_changed"] = False
    return series, pd.DataFrame(component_rows), median_differences


def validate_monthly_references(workbook: openpyxl.Workbook) -> tuple[int, int]:
    results = workbook[RESULT_SHEET]
    checks = matches = 0
    for column in range(3, results.max_column + 1):
        sheet_name = results.cell(1, column).value
        if not isinstance(sheet_name, str) or sheet_name not in workbook.sheetnames:
            continue
        month_sheet = workbook[sheet_name]
        for result_row, month_row in MONTH_SHEET_ROWS.items():
            result = results.cell(result_row, column).value
            source = month_sheet.cell(month_row, 3).value
            if not (_is_number(result) and _is_number(source)):
                continue
            checks += 1
            matches += abs(float(result) - float(source)) <= 1e-12
    return checks, matches


def validate_regsa_copy(workbook: openpyxl.Workbook, regsa_path: Path) -> tuple[int, int]:
    regsa = openpyxl.load_workbook(regsa_path, read_only=True, data_only=True)
    source = regsa["mom_sa"]
    target = workbook["ВВГУ_SА"] if "ВВГУ_SА" in workbook.sheetnames else workbook["ВВГУ_SA"]

    source_rows = [
        [cell.value for cell in row]
        for row in source.iter_rows(min_row=4, max_row=source.max_row, min_col=2, max_col=source.max_column)
    ]
    target_rows = [[cell.value for cell in row] for row in target.iter_rows()]
    transposed = list(map(list, zip(*source_rows)))

    checks = matches = 0
    for row_index in range(min(len(transposed), len(target_rows))):
        for column_index in range(min(len(transposed[row_index]), len(target_rows[row_index]))):
            expected = transposed[row_index][column_index]
            actual = target_rows[row_index][column_index]
            if _is_blank(expected) and _is_blank(actual):
                continue
            checks += 1
            if _is_number(expected) and _is_number(actual):
                matches += abs(float(expected) - float(actual)) <= 1e-12
            else:
                matches += expected == actual
    regsa.close()
    return checks, matches


def validate_mom_yoy_distinct(workbook: openpyxl.Workbook) -> tuple[int, int]:
    mom = _sheet_values(workbook, "ИПЦ исходный mom")
    yoy = _sheet_values(workbook, "ИПЦ исходный yoy")
    checks = identical = 0
    for row in range(min(len(mom), len(yoy))):
        for column in range(min(len(mom[row]), len(yoy[row]))):
            left, right = mom[row][column], yoy[row][column]
            if not (_is_number(left) and _is_number(right)):
                continue
            checks += 1
            identical += abs(float(left) - float(right)) <= 1e-12
    return checks, identical



def validate_database_yoy(workbook: openpyxl.Workbook, database_path: Path) -> tuple[int, int]:
    database = openpyxl.load_workbook(database_path, read_only=True, data_only=True, keep_links=True)
    database_rows = _sheet_values(database, database.active.title)
    database.close()
    if len(database_rows) < 5:
        return 0, 0

    dates = database_rows[3][2:]
    date_positions = {value: index for index, value in enumerate(dates) if value is not None}
    values_by_name = {
        row[1]: row[2:]
        for row in database_rows[4:]
        if len(row) > 2 and row[1] is not None
    }
    yoy = _sheet_values(workbook, "ИПЦ исходный yoy")
    if not yoy:
        return 0, 0
    names = yoy[0][1:]

    checks = matches = 0
    for row in yoy[1:]:
        if not row or row[0] not in date_positions:
            continue
        date_position = date_positions[row[0]]
        for column, name in enumerate(names, start=1):
            if name not in values_by_name or column >= len(row):
                continue
            database_values = values_by_name[name]
            if date_position >= len(database_values):
                continue
            actual, expected = row[column], database_values[date_position]
            if not (_is_number(actual) and _is_number(expected)):
                continue
            checks += 1
            matches += abs(float(actual) - float(expected)) <= 1e-12
    return checks, matches


def validate_official_formulas(workbook: openpyxl.Workbook) -> tuple[int, int, int]:
    results = workbook[RESULT_SHEET]
    checks = matches = 0
    result_rows = list(results.iter_rows(min_row=1, max_row=OFFICIAL_ROW, min_col=3, values_only=True))
    dates = result_rows[0]
    formulas = result_rows[OFFICIAL_ROW - 1]
    for offset, (date, formula) in enumerate(zip(dates, formulas, strict=False), start=3):
        if _parse_period(date) is None or not isinstance(formula, str) or not formula.startswith("="):
            continue
        letter = get_column_letter(offset)
        expected_references = [
            f"{letter}7:{letter}10",
            f"{letter}13:{letter}14",
            f"{letter}16:{letter}17",
            f"{letter}18",
            f"{letter}22:{letter}27",
            f"{letter}2",
        ]
        checks += 1
        matches += "MEDIAN(" in formula.upper() and all(reference in formula for reference in expected_references)

    random_count = 0
    target_rows = {137, 158, 180, 192}
    for sheet_name in workbook.sheetnames:
        if _parse_period(sheet_name) is None:
            continue
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=min(target_rows), max_row=max(target_rows), min_col=2, max_col=46, values_only=True),
            start=min(target_rows),
        ):
            if row_number not in target_rows:
                continue
            random_count += sum(isinstance(formula, str) and "RAND()" in formula.upper() for formula in row)
    return checks, matches, random_count


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def render_report(series: pd.DataFrame, validation: ValidationSummary) -> str:
    latest = series.tail(6)
    lines = [
        "# Воспроизведение утверждённой методики устойчивой инфляции",
        "",
        "Официальный показатель — строка 4 листа `Результаты`: медиана 16 показателей, перечисленных в формуле Excel. Месячный темп, 3MMA, диапазон оценок и медианная пара ниже являются прозрачным аналитическим слоем и не изменяют официальный ряд.",
        "",
        "## Последние шесть месяцев",
        "",
        "| месяц | официальный SAAR | SA м/м | 3MMA SAAR | IQR 16 оценок | медианная пара | пара изменилась |",
        "|---|---:|---:|---:|---:|---|:---:|",
    ]
    for row in latest.itertuples(index=False):
        pair = " ".join(f"{row.median_lower_label} / {row.median_upper_label}".replace("|", "/").split())
        lines.append(
            f"| {row.date[:7]} | {row.official_stable_inflation_saar:.2f}% | "
            f"{row.official_stable_inflation_mom_sa:.2f}% | {row.official_stable_inflation_3mma_saar:.2f}% | "
            f"{row.component_iqr_saar:.2f} п.п. | {pair} | {'да' if row.median_pair_changed else 'нет'} |"
        )
    lines.extend(
        [
            "",
            "## Как читать аналитический слой",
            "",
            "- `SA м/м` — месячный сезонно очищенный темп, обратным преобразованием полученный из официального SAAR.",
            "- `3MMA SAAR` — аннуализированное среднее трёх последних месячных темпов; это вспомогательный индикатор направления, а не новый официальный показатель.",
            "- `IQR 16 оценок` — межквартильный размах робастных оценок; чем он шире, тем выше методологическая неопределённость.",
            "- `Медианная пара` — 8-я и 9-я оценки после сортировки, среднее которых формирует официальную медиану.",
            "",
            "## Контроли",
            "",
            f"- Итоговая медиана: {validation.median_exact_matches}/{validation.official_months} месяцев совпали точно; максимальная разница {validation.median_max_abs_diff:.12g}.",
            f"- Формула официальной медианы: {validation.official_formula_exact_matches}/{validation.official_formula_checks} месячных формул содержат полный утверждённый набор ссылок.",
            f"- Ссылки на месячные листы: {validation.monthly_reference_exact_matches}/{validation.monthly_reference_checks} совпали точно.",
            f"- Перенос `RegSA.xlsx` → `ВВГУ_SA`: {validation.regsa_exact_matches}/{validation.regsa_checks} значимых ячеек совпали точно.",
            f"- `database.xlsx` → `ИПЦ исходный yoy`: {validation.database_yoy_exact_matches}/{validation.database_yoy_checks} общих числовых ячеек совпали точно.",
            f"- MoM/YoY: сопоставлено {validation.mom_yoy_comparable_cells} числовых ячеек; идентичны {validation.mom_yoy_identical_cells}, поэтому листы не являются копиями друг друга.",
            f"- Случайный tie-breaker: найдено {validation.random_tiebreaker_formula_count} формул с `RAND()`; это не блокирует воспроизведение сохранённого результата, но ограничивает детерминизм нового пересчёта Excel.",
            f"- Общий статус: {'PASS' if validation.passed else 'FAIL'}.",
            "",
            "## Ограничение",
            "",
            "Проверка воспроизводит формулы и сохранённые результаты утверждённой книги. Полный повтор шага 1 с переключением OLAP-меры `YoY2` → `MoM2` требует настольного Excel и доступной модели данных `database.xlsx`; LibreOffice/openpyxl не исполняют этот VBA/OLAP-контур.",
            "",
        ]
    )
    return "\n".join(lines)


def run(workbook_path: Path, database_path: Path, regsa_path: Path, output_path: Path) -> dict[str, Path]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True, keep_links=True)
    formula_workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False, keep_vba=True, keep_links=True)
    series, components, differences = extract_official_results(workbook)
    monthly_checks, monthly_matches = validate_monthly_references(workbook)
    regsa_checks, regsa_matches = validate_regsa_copy(workbook, regsa_path)
    database_checks, database_matches = validate_database_yoy(workbook, database_path)
    mom_yoy_checks, mom_yoy_identical = validate_mom_yoy_distinct(workbook)
    formula_checks, formula_matches, random_count = validate_official_formulas(formula_workbook)
    validation = ValidationSummary(
        official_months=len(series),
        median_exact_matches=sum(abs(value) <= 1e-12 for value in differences),
        median_max_abs_diff=max((abs(value) for value in differences), default=float("nan")),
        monthly_reference_checks=monthly_checks,
        monthly_reference_exact_matches=monthly_matches,
        regsa_checks=regsa_checks,
        regsa_exact_matches=regsa_matches,
        database_yoy_checks=database_checks,
        database_yoy_exact_matches=database_matches,
        mom_yoy_comparable_cells=mom_yoy_checks,
        mom_yoy_identical_cells=mom_yoy_identical,
        official_formula_checks=formula_checks,
        official_formula_exact_matches=formula_matches,
        random_tiebreaker_formula_count=random_count,
    )

    output_path.mkdir(parents=True, exist_ok=True)
    paths = {
        "series": output_path / "official_stable_inflation_series.csv",
        "components": output_path / "official_stable_inflation_components.csv",
        "validation": output_path / "official_validation.json",
        "source_manifest": output_path / "official_source_manifest.json",
        "report": output_path / "official_reproduction_report.md",
    }
    manifest = {
        name: {
            "path": str(path.resolve()),
            "size_bytes": path.stat().st_size,
            "sha256": _sha256(path),
        }
        for name, path in {
            "workbook": workbook_path,
            "database": database_path,
            "regsa": regsa_path,
        }.items()
    }
    series.to_csv(paths["series"], index=False)
    components.to_csv(paths["components"], index=False)
    paths["validation"].write_text(json.dumps(validation.__dict__ | {"passed": validation.passed}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    paths["source_manifest"].write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    paths["report"].write_text(render_report(series, validation), encoding="utf-8")
    workbook.close()
    formula_workbook.close()
    return paths


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Reproduce the approved stable-inflation workbook result.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--regsa", required=True, type=Path)
    parser.add_argument("--database", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args(argv)
    paths = run(args.workbook, args.database, args.regsa, args.output)
    for name, path in paths.items():
        print(f"{name}: {path}")
    validation = json.loads(paths["validation"].read_text(encoding="utf-8"))
    return 0 if validation["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
