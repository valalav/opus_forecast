#!/usr/bin/env python3
"""
Task 507: Extract KBR GDP Components

Extract quarterly GDP by economic sector for KBR.
Uses sectoral production indices as proxies for sectoral GDP.

OUTPUT: data/kbr_gdp_components.csv
"""

import pandas as pd
import openpyxl
from pathlib import Path
import re
import warnings

warnings.filterwarnings("ignore")


def parse_excel_structure(file_path):
    """Parse Excel file structure to find years, months, and KBR data."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active

        # Find month row (row 4)
        month_row = sheet[4] if sheet.max_row >= 4 else None

        # Find year row (row 3)
        year_row = sheet[3] if sheet.max_row >= 3 else None

        # Find KBR row
        kbr_row_idx = None
        for row_idx in range(1, min(200, sheet.max_row)):
            row = sheet[row_idx]
            if row[0] and row[0].value and "Кабардино" in str(row[0].value):
                kbr_row_idx = row_idx
                break

        if not kbr_row_idx or not month_row:
            return None, None, None

        kbr_row = sheet[kbr_row_idx]

        # Parse year and month structure
        data_points = []
        current_year = None
        month_names = []

        # Get month names from month_row
        for cell in month_row[1:]:  # Skip first column (region name)
            if cell.value:
                month_names.append(str(cell.value))

        # Parse data
        for col_idx in range(1, min(len(kbr_row), len(year_row))):
            year_cell = year_row[col_idx]
            month_cell = month_row[col_idx] if col_idx < len(month_row) else None
            value_cell = kbr_row[col_idx]

            # Update current year when we see a year marker
            if year_cell and year_cell.value:
                year_match = re.search(r"(\d{4})", str(year_cell.value))
                if year_match:
                    current_year = int(year_match.group(1))

            # Extract month data
            if month_cell and month_cell.value:
                month_name = str(month_cell.value)
                if value_cell.value and isinstance(value_cell.value, (int, float)):
                    value = float(value_cell.value)
                    if current_year:
                        data_points.append(
                            {"Year": current_year, "Month": month_name, "Value": value}
                        )

        return data_points, len(data_points) > 0, kbr_row_idx

    except Exception as e:
        print(f"  Error parsing Excel: {e}")
        import traceback

        traceback.print_exc()
        return None, False, None


def parse_sectoral_files():
    """Parse sectoral indices from sectoral files."""
    base_path = Path("./data/raw/info-stat")
    output_data = []

    # Map sectors to files
    sectors = [
        {
            "name": "Промышленность",
            "code": "A",
            "path": "01 промышленность/01-01 индекс промышленного производства.xlsx",
        },
        {
            "name": "Сельское хозяйство",
            "code": "B",
            "path": "02 сельское хозяйство/02-01 Индексы сельского хозяйства.xlsx",
        },
        {
            "name": "Строительство",
            "code": "F",
            "path": "03 строительство/03-01 объем работ выполненных по ВЭД Строительство.xlsx",
        },
        {
            "name": "Торговля",
            "code": "G",
            "path": "05 торговля/05-01 оборот розничной торговли.xlsx",
        },
    ]

    for sector in sectors:
        print(f"\n{'=' * 50}")
        print(f"Processing: {sector['name']}")
        print(f"{'=' * 50}")

        file_path = base_path / sector["path"]

        if not file_path.exists():
            print(f"  File not found: {file_path}")
            continue

        data_points, has_data, kbr_row = parse_excel_structure(file_path)

        if not has_data:
            print(f"  No data extracted")
            continue

        print(f"  Found {len(data_points)} data points at row {kbr_row}")

        # Convert to quarterly format
        for dp in data_points:
            # Determine quarter from month
            if dp["Month"] in ["январь", "февраль", "март"]:
                quarter = 1
            elif dp["Month"] in ["апрель", "май", "июнь"]:
                quarter = 2
            elif dp["Month"] in ["июль", "август", "сентябрь"]:
                quarter = 3
            else:
                quarter = 4

            output_data.append(
                {
                    "Date": f"{dp['Year']} Q{quarter}",
                    "Sector": sector["name"],
                    "Value": dp["Value"],
                }
            )

    return output_data


def calculate_yoy_growth(df):
    """Calculate Year-over-Year growth for each sector."""
    print("\nCalculating YoY growth...")

    df["Year"] = df["Date"].str.extract(r"(\d{4})").astype(int)
    df["Quarter"] = df["Date"].str.extract(r"Q(\d)").astype(float)

    result = []
    for sector in df["Sector"].unique():
        sector_data = df[df["Sector"] == sector].sort_values("Date").copy()
        sector_data = sector_data.reset_index(drop=True)

        # Calculate YoY growth (4 quarters = 1 year)
        values = sector_data["Value"].values
        yoy_growth = [0.0] * len(values)

        for i in range(4, len(values)):
            if values[i - 4] != 0:
                yoy_growth[i] = ((values[i] - values[i - 4]) / abs(values[i - 4])) * 100

        sector_data["YoY_growth"] = yoy_growth
        result.append(sector_data)

    return pd.concat(result, ignore_index=True)


def main():
    """Main extraction function."""
    print("=" * 60)
    print("Task 507: Extract KBR GDP Components")
    print("=" * 60)

    # Parse sectoral indices
    data = parse_sectoral_files()

    if not data:
        print("\nNo data extracted!")
        return None

    # Create DataFrame
    df = pd.DataFrame(data)

    print(f"\nTotal data points: {len(df)}")
    print(f"Sectors found: {df['Sector'].unique()}")
    print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")

    # Calculate YoY growth
    df = calculate_yoy_growth(df)

    # Add sector codes
    sector_code_map = {
        "Промышленность": "A",
        "Сельское хозяйство": "B",
        "Строительство": "F",
        "Торговля": "G",
    }
    df["Sector_Code"] = df["Sector"].apply(lambda x: sector_code_map.get(x, ""))

    # Select and reorder columns
    result = df[["Date", "Sector", "Sector_Code", "Value", "YoY_growth"]].copy()

    # Clean up NaN values
    result["YoY_growth"] = result["YoY_growth"].fillna(0.0)

    print(f"\nFinal output shape: {result.shape}")
    print(f"Date range: {result['Date'].min()} to {result['Date'].max()}")

    # Save to CSV
    output_path = Path("data/kbr_gdp_components.csv")
    output_path.parent.mkdir(exist_ok=True)
    result.to_csv(output_path, index=False)

    print(f"\n✓ Data saved to: {output_path}")

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY BY SECTOR")
    print("=" * 60)
    summary = result.groupby("Sector").agg(
        {"Value": ["count", "mean", "min", "max"], "YoY_growth": ["mean", "min", "max"]}
    )
    summary.columns = [
        "Count",
        "Mean_Value",
        "Min_Value",
        "Max_Value",
        "Mean_YoY",
        "Min_YoY",
        "Max_YoY",
    ]
    print(summary.round(2))

    print("\n" + "=" * 60)
    print("SAMPLE DATA (First 15 rows)")
    print("=" * 60)
    print(
        result[["Date", "Sector", "Value", "YoY_growth"]]
        .head(15)
        .to_string(index=False)
    )

    return result


if __name__ == "__main__":
    df = main()
