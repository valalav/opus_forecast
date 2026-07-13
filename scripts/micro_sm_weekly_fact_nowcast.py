from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from math import prod
from pathlib import Path

import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from sirena.data.weekly_bridge import load_semicolon_weekly_prices


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Blend Micro_SM item forecasts with weekly price facts for matched items."
    )
    parser.add_argument(
        "--forecast",
        default="data/external/micro_cpi_region_export/statsmodels_forecast.csv",
        help="Long Micro_SM statsmodels forecast CSV.",
    )
    parser.add_argument(
        "--actuals",
        default="data/external/micro_cpi_region_export/region_cpi_long.csv",
        help="Long CPI export with item weights and latest monthly facts.",
    )
    parser.add_argument(
        "--structure",
        default="data/items_structure.csv",
        help="Item hierarchy file; Item_type=5 rows are used as micro positions.",
    )
    parser.add_argument(
        "--weekly",
        default="data/Сравнение еженедельных цен_01.csv",
        help="Fresh semicolon weekly price file.",
    )
    parser.add_argument("--target-month", default="2026-05", help="Target month YYYY-MM.")
    parser.add_argument("--base-week", default="2026-04-27", help="Previous month-end week.")
    parser.add_argument("--fact-week", default="2026-05-25", help="Fact/latest week to use.")
    parser.add_argument(
        "--output-dir",
        default="archive/results/micro_sm_weekly_fact_20260608",
        help="Directory for CSV and report artifacts.",
    )
    return parser.parse_args()


def normalize_name(value: object) -> str:
    text = str(value).lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^а-яa-z0-9%*., -]+", "", text)
    return text.strip()


def pct_point_contribution(weight: float, index_value: float) -> float:
    return weight * (index_value - 100.0)


def weighted_index(frame: pd.DataFrame, value_col: str) -> float:
    valid = frame.dropna(subset=[value_col, "weight_vertical"])
    return float((valid[value_col] * valid["weight_vertical"]).sum() / valid["weight_vertical"].sum())


def build_micro_path(
    forecast: pd.DataFrame,
    micro_codes: set[str],
    micro_weights: pd.DataFrame,
    target_month: pd.Timestamp,
    replacements: pd.DataFrame,
) -> pd.DataFrame:
    replacement_values = replacements.set_index("item_code")["blended_value"].to_dict()
    rows = []
    for forecast_date, month_df in forecast[forecast["item_code"].isin(micro_codes)].groupby(
        "forecast_date"
    ):
        month = pd.Timestamp(forecast_date).to_period("M").to_timestamp()
        tmp = month_df.merge(
            micro_weights[["item_code", "weight_vertical"]],
            on="item_code",
            how="left",
        ).dropna(subset=["forecast_value", "weight_vertical"])
        tmp["hybrid_value"] = tmp["forecast_value"]
        if month == target_month:
            tmp["hybrid_value"] = tmp.apply(
                lambda row: replacement_values.get(row["item_code"], row["forecast_value"]),
                axis=1,
            )
        rows.append(
            {
                "date": month.date().isoformat(),
                "micro_weighted_index": weighted_index(tmp, "forecast_value"),
                "hybrid_index": weighted_index(tmp, "hybrid_value"),
                "micro_weight_sum": float(tmp["weight_vertical"].sum()),
                "weekly_replaced_items": int(len(replacement_values)) if month == target_month else 0,
            }
        )
    return pd.DataFrame(rows).sort_values("date")


def yoy_from_path(actual_indices: list[float], forecast_indices: list[float]) -> float:
    values = actual_indices + forecast_indices
    if len(values) != 12:
        return np.nan
    return round(prod(values) / (100**11), 2)


def main() -> None:
    args = parse_args()
    target_month = pd.Timestamp(args.target_month).to_period("M").to_timestamp()
    base_week = pd.Timestamp(args.base_week)
    fact_week = pd.Timestamp(args.fact_week)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    forecast = pd.read_csv(
        args.forecast,
        dtype={"item_code": "string", "item_rosstat": "string"},
        encoding="utf-8-sig",
    )
    forecast["forecast_date"] = pd.to_datetime(forecast["forecast_date"], errors="coerce")
    forecast["forecast_value"] = pd.to_numeric(forecast["forecast_value"], errors="coerce")

    actuals = pd.read_csv(
        args.actuals,
        dtype={"item_code": "string", "item_rosstat": "string"},
        encoding="utf-8-sig",
    )
    actuals["date"] = pd.to_datetime(actuals["date"], errors="coerce")
    actuals["weight_vertical"] = pd.to_numeric(actuals["weight_vertical"], errors="coerce")

    structure = pd.read_csv(args.structure, dtype={"Item_code": "string"})
    micro_codes = set(
        structure.loc[
            (structure["Item_type"] == 5) & (structure["Item_on"] == 1), "Item_code"
        ].astype(str)
    )

    latest_actual_month = actuals["date"].max().to_period("M").to_timestamp()
    latest_weights = (
        actuals.loc[actuals["date"].dt.to_period("M").dt.to_timestamp() == latest_actual_month]
        .drop_duplicates("item_code")
        [["item_code", "item_rosstat", "item_name", "weight_vertical"]]
    )
    micro_weights = latest_weights[latest_weights["item_code"].isin(micro_codes)].copy()

    month_forecast = forecast[
        forecast["forecast_date"].dt.to_period("M").dt.to_timestamp() == target_month
    ].copy()
    micro = month_forecast[month_forecast["item_code"].isin(micro_codes)].merge(
        micro_weights[["item_code", "weight_vertical"]],
        on="item_code",
        how="left",
        validate="one_to_one",
    )
    micro = micro.dropna(subset=["weight_vertical", "forecast_value"]).copy()
    micro["norm_name"] = micro["item_name"].map(normalize_name)

    weekly = load_semicolon_weekly_prices(args.weekly)
    weekly_pair = (
        weekly[weekly["date"].isin([base_week, fact_week])]
        .pivot_table(index="product_name", columns="date", values="price", aggfunc="last")
        .dropna()
        .reset_index()
    )
    weekly_pair["norm_name"] = weekly_pair["product_name"].map(normalize_name)
    weekly_pair["weekly_fact_index"] = (
        weekly_pair[fact_week] / weekly_pair[base_week] * 100.0
    )

    blended = micro.merge(
        weekly_pair[["product_name", "norm_name", base_week, fact_week, "weekly_fact_index"]],
        on="norm_name",
        how="left",
    )
    blended["used_weekly_fact"] = blended["weekly_fact_index"].notna()
    blended["blended_value"] = np.where(
        blended["used_weekly_fact"],
        blended["weekly_fact_index"],
        blended["forecast_value"],
    )
    blended["model_contribution_pp"] = blended.apply(
        lambda row: pct_point_contribution(row["weight_vertical"], row["forecast_value"]),
        axis=1,
    )
    blended["blended_contribution_pp"] = blended.apply(
        lambda row: pct_point_contribution(row["weight_vertical"], row["blended_value"]),
        axis=1,
    )
    blended["contribution_delta_pp"] = (
        blended["blended_contribution_pp"] - blended["model_contribution_pp"]
    )

    baseline_index = weighted_index(blended, "forecast_value")
    hybrid_index = weighted_index(blended, "blended_value")
    matched = blended[blended["used_weekly_fact"]].copy()

    direct_item1 = forecast[
        (forecast["item_code"].astype(str) == "1")
        & (forecast["forecast_date"].dt.to_period("M").dt.to_timestamp() == target_month)
    ]["forecast_value"]
    direct_item1_value = float(direct_item1.iloc[0]) if not direct_item1.empty else np.nan

    # YoY estimate uses official/near-official monthly path from the send-ready form when present.
    yoy_context = {}
    try:
        from openpyxl import load_workbook

        ws = load_workbook("assets/06_2026_02_Прогноз.xlsx", data_only=True)["Прогноз"]
        row_by_month = {
            ws.cell(row, 1).value.to_period("M").to_timestamp()
            if hasattr(ws.cell(row, 1).value, "to_period")
            else pd.Timestamp(ws.cell(row, 1).value).to_period("M").to_timestamp(): row
            for row in range(3, ws.max_row + 1)
            if ws.cell(row, 1).value is not None
        }
        target_row = row_by_month.get(target_month)
        if target_row:
            history = [float(ws.cell(r, 5).value) for r in range(target_row - 11, target_row)]
            yoy_context = {
                "hybrid_yoy_index": yoy_from_path(history, [hybrid_index]),
                "baseline_micro_yoy_index": yoy_from_path(history, [baseline_index]),
                "send_form_yoy_index": float(ws.cell(target_row, 6).value),
            }
    except Exception as exc:
        yoy_context = {"error": str(exc)}

    output_cols = [
        "item_code",
        "item_rosstat",
        "item_name",
        "weight_vertical",
        "forecast_value",
        "product_name",
        base_week,
        fact_week,
        "weekly_fact_index",
        "used_weekly_fact",
        "blended_value",
        "model_contribution_pp",
        "blended_contribution_pp",
        "contribution_delta_pp",
        "method",
        "status",
    ]
    # Timestamp column names are objects before CSV serialization.
    rename_dates = {base_week: str(base_week.date()), fact_week: str(fact_week.date())}
    blended_for_csv = blended.rename(columns=rename_dates)
    output_cols = [rename_dates.get(col, col) for col in output_cols]
    blended_for_csv[output_cols].to_csv(
        out_dir / "micro_sm_weekly_fact_items.csv",
        index=False,
        encoding="utf-8",
    )

    matched.sort_values("contribution_delta_pp", key=lambda s: s.abs(), ascending=False).rename(
        columns=rename_dates
    ).to_csv(out_dir / "micro_sm_weekly_fact_matched.csv", index=False, encoding="utf-8")

    path = build_micro_path(forecast, micro_codes, micro_weights, target_month, matched)
    path.to_csv(out_dir / "micro_sm_weekly_fact_path.csv", index=False, encoding="utf-8")

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "target_month": target_month.date().isoformat(),
        "base_week": base_week.date().isoformat(),
        "fact_week": fact_week.date().isoformat(),
        "forecast_file": args.forecast,
        "weekly_file": args.weekly,
        "micro_items": int(len(blended)),
        "micro_weight_sum": float(blended["weight_vertical"].sum()),
        "matched_items": int(matched.shape[0]),
        "matched_weight_sum": float(matched["weight_vertical"].sum()),
        "direct_item1_index": direct_item1_value,
        "baseline_micro_weighted_index": baseline_index,
        "hybrid_weekly_fact_index": hybrid_index,
        "baseline_micro_mom_pp": baseline_index - 100.0,
        "hybrid_weekly_fact_mom_pp": hybrid_index - 100.0,
        "hybrid_minus_baseline_pp": hybrid_index - baseline_index,
        **yoy_context,
    }
    (out_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    top_up = matched.sort_values("contribution_delta_pp", ascending=False).head(12)
    top_down = matched.sort_values("contribution_delta_pp", ascending=True).head(12)

    def table(frame: pd.DataFrame) -> str:
        rows = ["| Позиция | Вес | Модель | Факт weekly | Δ вклад, п.п. |", "|---|---:|---:|---:|---:|"]
        for _, row in frame.iterrows():
            rows.append(
                "| {name} | {w:.4f} | {model:.2f} | {fact:.2f} | {delta:+.4f} |".format(
                    name=row["item_name"],
                    w=row["weight_vertical"],
                    model=row["forecast_value"],
                    fact=row["weekly_fact_index"],
                    delta=row["contribution_delta_pp"],
                )
            )
        return "\n".join(rows)

    report = f"""# Micro_SM + weekly fact nowcast

Generated: {summary["generated_at"]}

## Setup

- Target month: `{summary["target_month"]}`
- Weekly fact window: `{summary["base_week"]}` -> `{summary["fact_week"]}`
- Micro items used: `{summary["micro_items"]}`
- Micro weight sum: `{summary["micro_weight_sum"]:.5f}`
- Matched weekly items: `{summary["matched_items"]}`
- Matched weight sum: `{summary["matched_weight_sum"]:.5f}`

## Headline Result

| Variant | Index | MoM, p.p. |
|---|---:|---:|
| Direct Micro_SM item_code=1 | {summary["direct_item1_index"]:.3f} | {summary["direct_item1_index"] - 100:.3f} |
| Micro_SM weighted micro aggregate | {summary["baseline_micro_weighted_index"]:.3f} | {summary["baseline_micro_mom_pp"]:.3f} |
| Micro_SM with weekly facts for matched items | {summary["hybrid_weekly_fact_index"]:.3f} | {summary["hybrid_weekly_fact_mom_pp"]:.3f} |

Hybrid minus weighted Micro_SM baseline: `{summary["hybrid_minus_baseline_pp"]:+.3f}` p.p.

## YoY Context

| Variant | YoY index |
|---|---:|
| Send form current | {summary.get("send_form_yoy_index", np.nan):.2f} |
| Weighted Micro_SM baseline for target month | {summary.get("baseline_micro_yoy_index", np.nan):.2f} |
| Hybrid weekly-fact target month | {summary.get("hybrid_yoy_index", np.nan):.2f} |

## Largest Upward Weekly Replacements

{table(top_up)}

## Largest Downward Weekly Replacements

{table(top_down)}

## 12-Month Micro Path

| Month | Micro_SM index | Hybrid index | Replaced items |
|---|---:|---:|---:|
{chr(10).join(f"| {row.date} | {row.micro_weighted_index:.3f} | {row.hybrid_index:.3f} | {int(row.weekly_replaced_items)} |" for row in path.itertuples())}

## Notes

- Weekly replacements are exact normalized-name matches only.
- The weekly fact is a price-level bridge from the last available April week to
  the last available May week, not an official monthly CPI fact.
- Unmatched micro items remain at their Micro_SM statsmodels forecast.
- Aggregation uses active `Item_type=5` micro positions and `weight_vertical`;
  their latest weight sum is `{summary["micro_weight_sum"]:.5f}`, so the weighted
  index is normalized by the included weight sum.
"""
    (out_dir / "report.md").write_text(report, encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
